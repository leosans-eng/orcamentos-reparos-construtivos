import tkinter as tk
from tkinter import ttk, filedialog
import pandas as pd
import json
import math
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import os
import unicodedata
from datetime import datetime

# ---------------------------- #
# CARREGAR DADOS JSON e CSV    #
# ---------------------------- #

with open("vicios_construtivos.json", "r", encoding="utf-8") as f:
    dados_json = json.load(f)

if isinstance(dados_json, list):
    dados_json = dados_json[0]

# CSV a ser lido. Trocar para a pasta sinapi/sinapi_processado no futuro
# Analisar forma de automatizar o arquivo a ser lido, para pegar o mais recente
sinapi = pd.read_csv("sinapi_precos.csv", dtype={"codigo": str})

sinapi.columns = sinapi.columns.str.strip().str.lower()

# ---------------------------- #
# JANELA PRINCIPAL             #
# ---------------------------- #

janela = tk.Tk()
janela.title("Orçamento de Reparos Construtivos - ORC")
janela.geometry("990x580+200+40")

# ---------------------------- #
# SCROLL DA JANELA             #
# ---------------------------- #

container = tk.Frame(janela)
container.pack(fill="both", expand=True)

canvas = tk.Canvas(container)
scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)

frame_principal = tk.Frame(canvas)

frame_principal.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas_window = canvas.create_window((0, 0), window=frame_principal, anchor="nw")

canvas.configure(yscrollcommand=scrollbar.set)

def atualizar_scroll(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

def _on_mousewheel(event):

    widget = janela.winfo_containing(event.x_root, event.y_root)

    if isinstance(widget, ttk.Combobox):
        return

    pos = canvas.yview()

    if event.delta > 0 and pos[0] <= 0:
        return

    canvas.yview_scroll(int(-1*(event.delta/120)), "units")

def ajustar_largura(event):
    canvas.itemconfig(canvas_window, width=event.width)

canvas.bind("<Configure>", ajustar_largura)

canvas.bind_all("<MouseWheel>", _on_mousewheel)

def ajustar_canvas(event):
    canvas.itemconfig(canvas_window, width=event.width)

canvas.bind("<Configure>", ajustar_canvas)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# ---------------------------- #
# FRAME DADOS DO ORÇAMENTO     #
# ---------------------------- #

frame_dados = tk.LabelFrame(frame_principal, text="Dados do Orçamento")
frame_dados.pack(fill="x", padx=10, pady=10)

# linha 1 - Proprietário / Autor
tk.Label(frame_dados, text="Autor(a):").grid(row=0, column=0, padx=5, pady=5, sticky="w")

entrada_proprietario = tk.Entry(frame_dados)
entrada_proprietario.grid(row=0, column=1, columnspan=3, sticky="ew", padx=5)

frame_dados.columnconfigure(1, weight=1)

tk.Label(frame_dados, text="Estado:").grid(row=0, column=4, padx=5)

estados = sorted(sinapi["estado"].dropna().unique())

combo_estado = ttk.Combobox(frame_dados, values=estados, width=6, state="readonly")
combo_estado.grid(row=0, column=5, padx=5)

def estado_alterado(event):

    if lista_anomalias:
        atualizar_tree()

combo_estado.bind("<<ComboboxSelected>>", estado_alterado)

def obter_estado():

    estado = combo_estado.get()

    if estado == "":
        mostrar_feedback("Selecione um Estado.", "red")
        raise ValueError

    return estado

# linha 2 - Estado, Aluguel e BDI

tk.Label(frame_dados, text="Aluguel (R$):").grid(row=1, column=2, padx=5)

entrada_aluguel = tk.Entry(frame_dados, width=10)
entrada_aluguel.grid(row=1, column=3, padx=5)
entrada_aluguel.insert(0, "1000")

tk.Label(frame_dados, text="BDI (%):").grid(row=1, column=4, padx=5)

entrada_bdi = tk.Entry(frame_dados, width=8)
entrada_bdi.grid(row=1, column=5, padx=5)
entrada_bdi.insert(0, "30,45")

# ---------------------------- #
# FRAME METRAGEM DOS CÔMODOS   #
# ---------------------------- #

frame_conteudo = tk.Frame(frame_principal)
frame_conteudo.pack(fill="both", expand=True)

frame_conteudo.columnconfigure(0, weight=0)
frame_conteudo.columnconfigure(1, weight=1)

frame_metragem = tk.LabelFrame(frame_conteudo, text="Metragem dos Cômodos")
frame_metragem.grid(row=0, column=0, sticky="nw", padx=10, pady=5)

frame_tabela = tk.Frame(frame_metragem)
frame_tabela.pack(pady=5)

tk.Label(frame_tabela, text="Cômodo", width=15, font=("Arial", 10, "bold")).grid(row=0, column=0)
tk.Label(frame_tabela, text="Piso (m²)", width=10, font=("Arial", 10, "bold")).grid(row=0, column=1)
tk.Label(frame_tabela, text="Rev. Arg. (m²)", width=12, font=("Arial", 10, "bold")).grid(row=0, column=2)
tk.Label(frame_tabela, text="Rev. Cer. (m²)", width=12, font=("Arial", 10, "bold")).grid(row=0, column=3)

lista_comodos = [
    "Sala",
    "Dormitório 1",
    "Dormitório 2",
    "Banheiro",
    "Cozinha",
    "Área de Serviço"
]

comodos_com_ceramica = [
    "Banheiro",
    "Cozinha",
    "Área de Serviço"
]

comodos = {}

for i, c in enumerate(lista_comodos, start=1):

    tk.Label(frame_tabela, text=c).grid(row=i, column=0)

    entrada_piso = tk.Entry(frame_tabela, width=10)
    entrada_piso.grid(row=i, column=1)

    entrada_rev_arg = tk.Entry(frame_tabela, width=10)
    entrada_rev_arg.grid(row=i, column=2)

    if c in comodos_com_ceramica:

        entrada_rev_cer = tk.Entry(frame_tabela, width=10)

    else:

        entrada_rev_cer = tk.Entry(frame_tabela, width=10, state="disabled")

    entrada_rev_cer.grid(row=i, column=3)

    comodos[c] = {
        "piso": entrada_piso,
        "rev_arg": entrada_rev_arg,
        "rev_cer": entrada_rev_cer
    }

# ---------------------------- #
# FRAME SELEÇÃO DE ANOMALIA    #
# ---------------------------- #

frame_anomalia = tk.LabelFrame(frame_conteudo, text="Selecionar Anomalia")
frame_anomalia.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

vicios = [
    "Desplacamento de pisos cerâmicos",
    "Desplacamento de azulejos",
    "Infiltração pela esquadria"
]

combo_vicio = ttk.Combobox(frame_anomalia, width=40, values=vicios, state="readonly")
combo_vicio.pack(pady=5)

comodos_bloqueados_por_vicio = {
    "Desplacamento de azulejos": {"Sala", "Dormitório 1", "Dormitório 2"}
}

# ---------------------------- #
# CHECKBOXES DE CÔMODOS        #
# ---------------------------- #

tk.Label(frame_anomalia, text="Cômodos afetados:").pack()

frame_check = tk.Frame(frame_anomalia)
frame_check.pack(pady=5)

linhas = [
    ["Sala", "Banheiro"],
    ["Dormitório 1", "Cozinha"],
    ["Dormitório 2", "Área de Serviço"]
]

checkbox_comodos = {}

for r, linha in enumerate(linhas):

    for c, comodo in enumerate(linha):

        var = tk.BooleanVar()

        chk = tk.Checkbutton(frame_check, text=comodo, variable=var)
        chk.grid(row=r, column=c, sticky="w", padx=10)

        checkbox_comodos[comodo] = {
            "var": var,
            "widget": chk
        }

def atualizar_checkboxes_por_vicio(event=None):

    vicio_selecionado = combo_vicio.get()

    comodos_bloqueados = comodos_bloqueados_por_vicio.get(vicio_selecionado, set())

    for comodo, dados in checkbox_comodos.items():

        var = dados["var"]
        chk = dados["widget"]

        if comodo in comodos_bloqueados:
            var.set(False)
            chk.config(state="disabled")
        else:
            chk.config(state="normal")

combo_vicio.bind("<<ComboboxSelected>>", atualizar_checkboxes_por_vicio)

# ---------------------------- #
# FEEDBACK VISUAL              #
# ---------------------------- #

feedback_timer = None

def mostrar_feedback(mensagem, cor="red", temporario=True):

    global feedback_timer

    feedback_label.config(text=mensagem, fg=cor)

    if feedback_timer is not None:
        janela.after_cancel(feedback_timer)

    if temporario:
        feedback_timer = janela.after(
            3000,
            lambda: feedback_label.config(text="")
    )

# ---------------------------- #
# LISTA DE ANOMALIAS           #
# ---------------------------- #

frame_lista = tk.LabelFrame(frame_conteudo, text="Anomalias adicionadas")
frame_lista.grid(row=0, column=1, rowspan=2, sticky="n", padx=10, pady=5)

lista_anomalias = []

frame_listbox = tk.Frame(frame_lista)
frame_listbox.pack(fill="both", expand=True)

tree_anomalias = ttk.Treeview(
    frame_listbox,
    columns=("subtotal"),
    show="tree headings",
    height=11
)

tree_anomalias.heading("#0", text="Anomalia")
tree_anomalias.heading("subtotal", text="Subtotal")

tree_anomalias.column("#0", width=350)
tree_anomalias.column("subtotal", width=100, anchor="e")

tree_anomalias.pack(side="left", fill="both", expand=True)

scroll_lista = ttk.Scrollbar(
    frame_listbox,
    orient="vertical",
    command=tree_anomalias.yview
)

tree_anomalias.configure(yscrollcommand=scroll_lista.set)

scroll_lista.pack(side="right", fill="y")

# ---------------------------- #
# FUNÇÃO ADICIONAR ANOMALIA    #
# ---------------------------- #

def adicionar_anomalia():

    vicio = combo_vicio.get()

    if not vicio:
        mostrar_feedback("Selecione uma anomalia.", "red")
        return

    comodos_afetados = [
        c for c in lista_comodos if checkbox_comodos[c]["var"].get()
    ]

    if not comodos_afetados:
        mostrar_feedback("Selecione ao menos um cômodo.", "red")
        return
    
    # procurar se a anomalia já foi adicionada
    for item in lista_anomalias:

        if item["vicio"] == vicio:

            novos = []

            for c in comodos_afetados:

                if c not in item["comodos"]:
                    item["comodos"].append(c)
                    novos.append(c)

            if novos:

                atualizar_tree()

                mostrar_feedback(
                    f"Cômodo(s) adicionado(s): {', '.join(novos)}",
                    "orange"
                )

            else:

                mostrar_feedback(
                    "Esta anomalia já está cadastrada nesse(s) cômodo(s).",
                    "orange red"
                )

            for dados in checkbox_comodos.values():
                dados["var"].set(False)

            return
        
    # criar nova anomalia
    item = {
        "vicio": vicio,
        "comodos": comodos_afetados
    }

    lista_anomalias.append(item)

    atualizar_tree()

    for dados in checkbox_comodos.values():
        dados["var"].set(False)

    mostrar_feedback("Anomalia adicionada com sucesso.", "green")

def atualizar_tree():

    tree_anomalias.delete(*tree_anomalias.get_children())

    for item in lista_anomalias:

        subtotal = calcular_subtotal_anomalia(item)

        subtotal_str = f"R$ {subtotal:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        pai = tree_anomalias.insert(
            "",
            "end",
            text=item["vicio"],
            values=(subtotal_str,),
            open=True
        )

        for comodo in item["comodos"]:

            tree_anomalias.insert(
                pai,
                "end",
                text=comodo,
                values=("",)
            )
# ---------------------------- #
# FUNÇÃO REMOVER ANOMALIA      #
# ---------------------------- #

def remover_anomalia():

    selecionado = tree_anomalias.selection()

    if not selecionado:
        mostrar_feedback("Selecione uma anomalia para remover.", "red")
        return

    item_id = selecionado[0]

    pai = tree_anomalias.parent(item_id)

    # SE SELECIONAR APENAS A ANOMALIA
    if pai == "":

        indice = tree_anomalias.index(item_id)
        del lista_anomalias[indice]
        atualizar_tree()
        mostrar_feedback("Anomalia removida.", "orange red")

    # SE SELECIONAR UM CÔMODO
    else:

        indice_anomalia = tree_anomalias.index(pai)
        comodo = tree_anomalias.item(item_id)["text"]
        lista_anomalias[indice_anomalia]["comodos"].remove(comodo)

        # PARA TIRAR A ANOMALIA CASO NÃO SOBRE NENHUM CÔMODO
        if not lista_anomalias[indice_anomalia]["comodos"]:
            del lista_anomalias[indice_anomalia]

        atualizar_tree()

        mostrar_feedback(f"Cômodo removido: {comodo}", "orange")

def remover_todas_anomalias():

    if not lista_anomalias:
        mostrar_feedback("Não há anomalias para remover.", "orange")
        return

    lista_anomalias.clear()
    atualizar_tree()
    mostrar_feedback("Todas as anomalias foram removidas.", "orange red")
    
# ---------------------------- #
# BOTÕES                       #
# ---------------------------- #

tk.Button(
    frame_anomalia,
    text="Adicionar Anomalia",
    command=adicionar_anomalia
).pack(pady=10)

tk.Button(
    frame_lista,
    text="Remover Anomalia Selecionada",
    command=remover_anomalia
).pack(pady=5)

tk.Button(
    frame_lista,
    text="Remover Todas as Anomalias",
    underline=8,
    command=remover_todas_anomalias
).pack(pady=5)

# ---------------------------- #
# ÁREA DE FEEDBACK             #
# ---------------------------- #

frame_feedback = tk.Frame(frame_principal)
frame_feedback.pack(pady=(5,0))

feedback_label = tk.Label(
    frame_feedback,
    text="",
    font=("Arial", 10, "bold"),
    fg="#a67c00"
)

feedback_label.pack()

# ---------------------------- #
# FUNÇÃO CALCULAR QUANTIDADE   #
# ---------------------------- #

def ler_float(entry):

    valor = entry.get().strip()

    if valor == "":
        return 0

    valor = valor.replace(",", ".")

    try:
        return float(valor)
    except:
        raise ValueError
    
def calcular_quantidade(etapa, medidas):

    tipo = etapa["tipo_calculo"]

    if tipo == "area_piso":
        return medidas["piso"] * etapa.get("coeficiente", 1)

    if tipo == "area_rev_arg":
        return medidas["rev_arg"] * etapa.get("coeficiente", 1)

    if tipo == "area_rev_cer":
        return medidas["rev_cer"] * etapa.get("coeficiente", 1)

    if tipo == "perimetro":
        return math.sqrt(medidas["piso"]) * 4 * etapa.get("coeficiente", 1)

    if tipo == "por_comodo":
        return etapa["quantidade"]

    if tipo == "fixo":
        return etapa["quantidade"]

    return 0

def calcular_subtotal_anomalia(item):

    total = 0

    nome_anomalia = item["vicio"]
    etapas = dados_json["anomalias"][nome_anomalia]["etapas"]

    for comodo in item["comodos"]:

        piso = ler_float(comodos[comodo]["piso"])
        arg = ler_float(comodos[comodo]["rev_arg"])

        if comodos[comodo]["rev_cer"].cget("state") != "disabled":
            cer = ler_float(comodos[comodo]["rev_cer"])
        else:
            cer = 0

        medidas = {
            "piso": piso,
            "rev_arg": arg,
            "rev_cer": cer
        }

        for etapa in etapas:

            quantidade = calcular_quantidade(etapa, medidas)

            codigo = str(etapa["codigo_sinapi"])

            estado = combo_estado.get()

            linha_sinapi = sinapi[
                (sinapi["codigo"] == codigo) &
                (sinapi["estado"] == estado)
            ]

            if not linha_sinapi.empty:
                valor = linha_sinapi.iloc[0]["custo"]
            else:
                valor = 0

            total += quantidade * valor

    return total

# ---------------------------- #
# AUXILIAR DE NOME DE ARQUIVO  #
# ---------------------------- #

def nome_arquivo(texto):

    texto = texto.strip()

    if not texto:
        return "sem_proprietario"

    caracteres_invalidos = '<>:"/\\|?*'

    for caractere in caracteres_invalidos:
        texto = texto.replace(caractere, "_")

    texto = "_".join(texto.split())

    return texto[:80]

def normalizar_texto(texto):
    if texto is None:
        return ""
    texto = str(texto).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.upper()

# ---------------------------- #
# GERAR ORÇAMENTO              #
# ---------------------------- #

def gerar_orcamento():

    #Apenas para diagnosticar alguns bugs. Excluir essa print depois de gerar o primeiro orçamento completo
    print("LISTA_ANOMALIAS:", lista_anomalias)

    if not lista_anomalias:
        mostrar_feedback(
            "Adicione ao menos uma anomalia para gerar o orçamento.",
            "red"
        )
        return

    linhas = []

    for item in lista_anomalias:

        if not isinstance(item, dict):
            print("Item inválido na lista:", item)
            continue

        nome_anomalia = item["vicio"]
        comodos_afetados = item["comodos"]

        print("Anomalia selecionada:", nome_anomalia)
        print("Chaves disponíveis no JSON:", dados_json["anomalias"].keys())

        etapas = dados_json["anomalias"][nome_anomalia]["etapas"]

        for comodo in comodos_afetados:

            try:

                piso = ler_float(comodos[comodo]["piso"])
                arg = ler_float(comodos[comodo]["rev_arg"])

                if comodos[comodo]["rev_cer"].cget("state") != "disabled":
                    cer = ler_float(comodos[comodo]["rev_cer"])
                else:
                    cer = 0

            except Exception as e:

                print("ERRO:", e)

                mostrar_feedback(
                    f"Erro ao ler medidas do cômodo {comodo}",
                    "red"
                )

                return

            medidas = {
                "piso": piso,
                "rev_arg": arg,
                "rev_cer": cer
            }

            for ordem, etapa in enumerate(etapas):

                quantidade = calcular_quantidade(etapa, medidas)

                codigo = str(etapa["codigo_sinapi"])

                estado = obter_estado()

                linha_sinapi = sinapi[
                    (sinapi["codigo"] == codigo) &
                    (sinapi["estado"] == estado)
                ]

                if not linha_sinapi.empty:

                    descricao = linha_sinapi.iloc[0]["descricao"]
                    valor = linha_sinapi.iloc[0]["custo"]

                else:

                    descricao = "Código não encontrado"
                    valor = 0

                total = quantidade * valor

                linhas.append({
                    "Anomalia": nome_anomalia,
                    "Ordem": ordem,
                    "Código SINAPI": codigo,
                    "Descrição do item": descricao,
                    "Unid.": etapa["unidade"],
                    "Qtd.": round(quantidade, 2),
                    "Valor Unit.": valor,
                    "Total s/ BDI": round(total, 2)
                })

    df = pd.DataFrame(linhas)

    df = df.sort_values(["Anomalia", "Ordem"])

    df = df.groupby(
        ["Anomalia", "Código SINAPI", "Descrição do item", "Unid.", "Valor Unit.", "Ordem"],
        as_index=False,
        sort=False
    ).agg({
        "Qtd.": "sum",
        "Total s/ BDI": "sum"
    })

    total_geral = df["Total s/ BDI"].sum()

    df = df.sort_values(["Anomalia", "Ordem"])

    linhas_final = []

    df = df.drop(columns=["Ordem"])

    for anomalia, grupo in df.groupby("Anomalia"):

        subtotal = grupo["Total s/ BDI"].sum()

        # subtítulo da anomalia
        linhas_final.append({
            "Código SINAPI": "",
            "Descrição do item": anomalia.upper(),
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": subtotal
        })

        # itens SINAPI
        for _, row in grupo.iterrows():

            linhas_final.append({
                "Código SINAPI": row["Código SINAPI"],
                "Descrição do item": row["Descrição do item"],
                "Unid.": row["Unid."],
                "Qtd.": row["Qtd."],
                "Valor Unit.": row["Valor Unit."],
                "Total s/ BDI": row["Total s/ BDI"]
            })

    df = pd.DataFrame(linhas_final)

    df["Total s/ BDI"] = pd.to_numeric(df["Total s/ BDI"], errors="coerce").fillna(0)

    linha_total = pd.DataFrame([{
        "Código SINAPI": "",
        "Descrição do item": "Total sem BDI",
        "Unid.": "",
        "Qtd.": "",
        "Valor Unit.": "",
        "Total s/ BDI": round(total_geral, 2)
    }])

    df = pd.concat([df, linha_total], ignore_index=True)

    # ------------ #
    # CALCULAR BDI #
    # ------------ #

    bdi_str = entrada_bdi.get().replace(",", ".")

    try:
        bdi = float(bdi_str) / 100
    except:
        mostrar_feedback("BDI inválido.", "red")
        return

    valor_bdi = total_geral * bdi

    linha_bdi = pd.DataFrame([{
        "Código SINAPI": "",
        "Descrição do item": f"Total do BDI ({entrada_bdi.get()}%)",
        "Unid.": "",
        "Qtd.": "",
        "Valor Unit.": "",
        "Total s/ BDI": round(valor_bdi, 2)
    }])

    df = pd.concat([df, linha_bdi], ignore_index=True)

    # ------- #
    # ALUGUEL #
    # ------- #

    aluguel_str = entrada_aluguel.get().replace(",", ".")

    try:
        aluguel = float(aluguel_str)
    except:
        mostrar_feedback("Valor de aluguel inválido.", "red")
        return

    linha_aluguel = pd.DataFrame([{
        "Código SINAPI": "",
        "Descrição do item": "Aluguel (1 mês)",
        "Unid.": "",
        "Qtd.": "",
        "Valor Unit.": "",
        "Total s/ BDI": round(aluguel, 2)
    }])

    df = pd.concat([df, linha_aluguel], ignore_index=True)

    # ----------- #
    # TOTAL FINAL #
    # ----------- #

    total_final = total_geral + valor_bdi + aluguel

    linha_total_final = pd.DataFrame([{
        "Código SINAPI": "",
        "Descrição do item": "TOTAL GERAL",
        "Unid.": "",
        "Qtd.": "",
        "Valor Unit.": "",
        "Total s/ BDI": round(total_final, 2)
    }])

    df = pd.concat([df, linha_total_final], ignore_index=True)

    df = df[["Código SINAPI", "Descrição do item", "Unid.", "Qtd.", "Valor Unit.", "Total s/ BDI"]]

    nome_proprietario = nome_arquivo(entrada_proprietario.get())
    nome_base = f"orcamento_reparos_{nome_proprietario}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    pasta_downloads = os.path.join(os.path.expanduser("~"), "Downloads")

    if not os.path.isdir(pasta_downloads):
        pasta_downloads = os.getcwd()

    arquivo = filedialog.asksaveasfilename(
        title="Salvar orçamento como:",
        initialdir=pasta_downloads,
        initialfile=nome_base,
        defaultextension=".xlsx",
        filetypes=[("Planilha Excel", "*.xlsx")]
    )

    if not arquivo:
        mostrar_feedback("Geração de orçamento cancelada.", "orange", temporario=False)
        return

    try:
        df.to_excel(arquivo, index=False)
    except PermissionError:
        mostrar_feedback(
            "Feche a planilha antes de gerar novamente.",
            "red",
            temporario=False
        )
        return

    mostrar_feedback("Orçamento gerado com sucesso!", "dark green", temporario=False)
    
# ---------------------------- #
# FORMATAR PLANILHA            #
# ---------------------------- #

    wb = load_workbook(arquivo)
    ws = wb.active

    fonte_cabecalho = Font(bold=True)

    fundo_cabecalho = PatternFill(
        start_color="006699",
        end_color="006699",
        fill_type="solid"
    )

    fundo_anomalia = PatternFill(
        start_color='D0CECE',
        end_color='D0CECE',
        fill_type='solid'
    )
    
    fundo_totais = PatternFill(
        start_color="F2F2F2",
        end_color="F2F2F2",
        fill_type="solid"
    )

    fundo_total_final = PatternFill(
        start_color="006699",
        end_color="006699",
        fill_type="solid"
    )
    
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6):
        for cell in row:
            cell.border = borda
    
    # Linha de título logo acima do cabeçalho original
    ws.insert_rows(1)
    nome_titulo = entrada_proprietario.get().strip() or ""
    ws.merge_cells("A1:F1")
    ws["A1"] = f"ORÇAMENTO DE REPAROS DE VÍCIOS CONSTRUTIVOS - {nome_titulo.upper()}"
    ws.row_dimensions[1].height = 24.75

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    # larguras das colunas
    ws.column_dimensions["A"].width = 7.24
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 5.96

    # largura da coluna de Qtd.
    max_length = 0

    for col in ["D"]:

        for cell in ws[col]:

            if cell.value is not None:

                comprimento = len(str(cell.value))

                if comprimento > max_length:
                    max_length = comprimento

    ws.column_dimensions["D"].width = max_length + 1

    # largura da coluna de Total s/ BDI
    max_length = 0

    for col in ["F"]:

        for cell in ws[col]:

            if cell.value is not None:

                comprimento = len(str(cell.value))

                if comprimento > max_length:
                    max_length = comprimento

    ws.column_dimensions["F"].width = max_length

    # centralizar e quebrar texto do cabeçalho
    # título (linha 1)
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = fundo_cabecalho
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # cabeçalho da tabela (linha 2)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fundo_cabecalho
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
    
    colunas_centro = ["A", "C", "D", "E", "F"]

    for col in colunas_centro:

        for cell in ws[col]:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):

        for cell in row:

            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal if cell.alignment else "center",
                vertical="center",
                wrap_text=True
            )
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):

        codigo = row[0].value
        descricao = row[1].value

        # --------------------------- #
        # TOTAL SEM BDI, BDI, ALUGUEL #
        # --------------------------- #
        descricao_norm = normalizar_texto(descricao)

        if descricao_norm and (
            descricao_norm == "TOTAL SEM BDI" or
            descricao_norm.startswith("TOTAL DO BDI") or
            descricao_norm == "ALUGUEL (1 MES)"
        ):
            linha_idx = row[0].row
            descricao_formatada = str(descricao).strip()

            if descricao_norm == "TOTAL SEM BDI":
                descricao_formatada = "Total sem BDI"
            elif descricao_norm == "ALUGUEL (1 MES)":
                descricao_formatada = "Aluguel (1 mês)"
            elif descricao_norm.startswith("TOTAL DO BDI"):
                descricao_formatada = f"Total do BDI ({entrada_bdi.get()}%)"

            # O texto original está na coluna B. Após mesclar, precisa ficar em A
            ws.cell(row=linha_idx, column=1, value=descricao_formatada)
            for col_idx in range(2, 6):
                ws.cell(row=linha_idx, column=col_idx, value=None)

            # Mesclar rótulo em A:E, como no modelo de referência.
            ws.merge_cells(start_row=linha_idx, start_column=1, end_row=linha_idx, end_column=5)

            for col_idx in range(1, 7):
                cell = ws.cell(row=linha_idx, column=col_idx)
                cell.fill = fundo_totais
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(
                    horizontal="right" if col_idx == 1 else "center",
                    vertical="center",
                    wrap_text=True
                )

        # ----------- #
        # TOTAL FINAL #
        # ----------- #
        elif descricao == "TOTAL GERAL":
            linha_idx = row[0].row
            descricao_formatada = "Total do Orçamento"

            ws.cell(row=linha_idx, column=1, value=descricao_formatada)
            for col_idx in range(2, 6):
                ws.cell(row=linha_idx, column=col_idx, value=None)

            ws.merge_cells(start_row=linha_idx, start_column=1, end_row=linha_idx, end_column=5)

            for col_idx in range(1, 7):
                cell = ws.cell(row=linha_idx, column=col_idx)
                cell.fill = fundo_total_final
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.alignment = Alignment(
                    horizontal="right" if col_idx == 1 else "center",
                    vertical="center",
                    wrap_text=True
                )

        # --------------------- #
        # SUBTÍTULO DE ANOMALIA #
        # --------------------- #
        elif (codigo in ("", None)) and descricao not in ("", None):

            for cell in row:
                cell.fill = fundo_anomalia
                cell.font = Font(bold=True, size=12)


    wb.save(arquivo)
    os.startfile(arquivo)

# ---------------------------- #
# BOTÃO GERAR ORÇAMENTO        #
# ---------------------------- #

botao_gerar = tk.Button(
    frame_principal,
    text="Gerar Orçamento",
    font=("Arial", 11, "bold"),
    padx=12,
    pady=6,
    #   # cores do botão
    #   bg="#006699",
    #   fg="white",
    #   activebackground="#00557a",
    #   activeforeground="white",
    #   relief="flat",
    #   bd=0,
    #   cursor="hand2",

    command=gerar_orcamento
)

botao_gerar.pack(pady=(5,10))

entrada_proprietario.focus()

janela.mainloop()