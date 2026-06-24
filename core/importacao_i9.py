"""Importação de planilhas sintéticas exportadas pelo sistema i9."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from core.composicoes_proprias_storage import obter_por_codigo
from core.formatador_sinapi.comum import (
    encontrar_linha_dados_start,
    extrair_bdi_rotulo,
    extrair_estado_sinapi,
    extrair_nome_obra,
    planilha_ativa,
)
from core.formatador_sinapi.entrada import validar_caminho_planilha
from core.orcamento_customizado import BDI_PADRAO, OrcamentoCustomizado


@dataclass
class ResultadoImportacaoI9:
    orcamento: OrcamentoCustomizado
    avisos: list[str] = field(default_factory=list)
    grupos_importados: int = 0
    itens_importados: int = 0


def _texto_celula(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _parse_float(valor) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    return float(texto)


def _parse_bdi_percentual(rotulo: str) -> float:
    try:
        return float(rotulo.replace(",", "."))
    except ValueError:
        return BDI_PADRAO


def _eh_linha_totais(ws, linha: int) -> bool:
    for coluna in (1, 9):
        texto = _texto_celula(ws.cell(row=linha, column=coluna).value)
        if texto.startswith("Total"):
            return True
    return False


def _eh_linha_grupo(ws, linha: int) -> bool:
    codigo = _texto_celula(ws.cell(row=linha, column=4).value)
    descricao = _texto_celula(ws.cell(row=linha, column=5).value)
    banco = _texto_celula(ws.cell(row=linha, column=3).value)
    if codigo:
        return False
    if not descricao:
        return False
    if banco:
        return False
    return True


def _nome_orcamento_importado(ws, caminho: Path) -> str:
    nome_obra = extrair_nome_obra(ws).strip()
    if nome_obra:
        return nome_obra
    return f"Importado — {caminho.stem}"


def importar_planilha_i9(caminho: str, catalogo=None) -> ResultadoImportacaoI9:
    """
    Lê uma planilha sintética i9 e monta um OrcamentoCustomizado editável.

    Itens SINAPI (composições ou insumos) viram itens SINAPI no orçamento.
    Itens com banco Próprio são vinculados ao catálogo de composições próprias pelo código.
    """
    arquivo = validar_caminho_planilha(caminho)
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = planilha_ativa(wb)

    nome = _nome_orcamento_importado(ws, arquivo)
    bdi = _parse_bdi_percentual(extrair_bdi_rotulo(ws))
    estado = extrair_estado_sinapi(ws)

    orcamento = OrcamentoCustomizado(nome=nome, bdi_percent=bdi)
    orcamento.definir_estado_referencia(estado)

    resultado = ResultadoImportacaoI9(orcamento=orcamento)
    linha_inicio = encontrar_linha_dados_start(ws)
    grupo_atual_id = None

    for linha in range(linha_inicio, ws.max_row + 1):
        if _eh_linha_totais(ws, linha):
            break

        item = _texto_celula(ws.cell(row=linha, column=1).value)
        banco = _texto_celula(ws.cell(row=linha, column=3).value)
        codigo = _texto_celula(ws.cell(row=linha, column=4).value)
        descricao = _texto_celula(ws.cell(row=linha, column=5).value)
        unidade = _texto_celula(ws.cell(row=linha, column=6).value)
        quantidade = _parse_float(ws.cell(row=linha, column=7).value)
        preco_sem_bdi = _parse_float(ws.cell(row=linha, column=8).value)

        if not any((item, banco, codigo, descricao, unidade)) and quantidade == 0:
            continue

        if _eh_linha_grupo(ws, linha):
            grupo_atual_id = orcamento.adicionar_grupo(descricao)
            resultado.grupos_importados += 1
            continue

        if not codigo:
            continue

        if grupo_atual_id is None:
            grupo_atual_id = orcamento.adicionar_grupo("GERAL")
            resultado.grupos_importados += 1

        if quantidade <= 0:
            resultado.avisos.append(
                f"Linha {linha}: quantidade inválida para o item {codigo}; ignorado."
            )
            continue

        banco_normalizado = banco.lower().replace("ó", "o")
        if banco_normalizado == "proprio":
            dados_catalogo = {"composicoes": catalogo} if catalogo is not None else None
            composicao = obter_por_codigo(codigo, dados_catalogo)
            if composicao is None:
                rotulo = descricao or codigo
                resultado.avisos.append(
                    f"Composição própria não encontrada no catálogo: {codigo} — {rotulo}"
                )
                continue
            orcamento.adicionar_composicao_propria(
                grupo_atual_id,
                composicao["id"],
                composicao.get("codigo", codigo),
                composicao.get("nome", descricao),
                composicao.get("unidade", unidade) or unidade,
                quantidade,
            )
        else:
            if preco_sem_bdi <= 0:
                resultado.avisos.append(
                    f"Linha {linha}: preço unitário ausente para SINAPI {codigo}; "
                    "item importado com custo zero."
                )
            orcamento.adicionar_item_sinapi(
                grupo_atual_id,
                codigo,
                descricao,
                unidade,
                preco_sem_bdi,
                quantidade,
                estado,
            )

        resultado.itens_importados += 1

    if resultado.grupos_importados == 0 and resultado.itens_importados == 0:
        raise ValueError(
            "Nenhuma etapa ou item reconhecido na planilha.\n"
            "Verifique se o arquivo segue o layout de Planilha Sintética do i9."
        )

    return resultado
