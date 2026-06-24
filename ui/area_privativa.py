import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import math
import os
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

from core.app_state import ALTURA_TREE_MIN, LARGURA_JANELA_PADRAO, NOMES_GRUPOS_REPARO
from ui.widgets import criar_barra_modulo


def criar_area_privativa(parent, ctx, on_voltar):
    """Monta o modulo de orcamento para area privativa."""
    root = parent.winfo_toplevel()
    lista_anomalias = []
    _feedback_timer = None

    wrapper = tk.Frame(parent)

    criar_barra_modulo(wrapper, "Área Privativa", on_voltar)

    # SCROLL DA JANELA             #
    # ---------------------------- #
    container = tk.Frame(wrapper)
    container.pack(fill="both", expand=True)

    canvas = tk.Canvas(container, highlightthickness=0, bg="#ececec")
    scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)

    frame_principal = tk.Frame(canvas, bg=canvas["bg"])

    def _atualizar_scrollregion(_event=None):
        canvas.configure(scrollregion=canvas.bbox("all"))

    frame_principal.bind("<Configure>", _atualizar_scrollregion)

    canvas_window = canvas.create_window((0, 0), window=frame_principal, anchor="nw")

    canvas.configure(yscrollcommand=scrollbar.set)

    def _conteudo_cabe_no_canvas():
        bbox = canvas.bbox("all")
        if not bbox:
            return True
        return (bbox[3] - bbox[1]) <= canvas.winfo_height()

    def _on_mousewheel(event):

        widget = root.winfo_containing(event.x_root, event.y_root)

        if isinstance(widget, ttk.Combobox):
            return

        if _conteudo_cabe_no_canvas():
            return

        pos = canvas.yview()

        if event.delta > 0 and pos[0] <= 0:
            return

        canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def ajustar_layout_canvas(event):
        largura_canvas = event.width
        altura_canvas = event.height
        largura_conteudo = min(largura_canvas, LARGURA_JANELA_PADRAO)
        offset_x = max(0, (largura_canvas - largura_conteudo) // 2)

        canvas.itemconfig(canvas_window, width=largura_conteudo)
        canvas.coords(canvas_window, offset_x, 0)

        if not getattr(ajustar_layout_canvas, "_pronto", False):
            _atualizar_scrollregion()
            return

        frame_conteudo.pack_propagate(True)
        frame_conteudo.configure(height=0)
        frame_principal.update_idletasks()
        altura_natural = frame_principal.winfo_reqheight()

        if altura_canvas > altura_natural + 8:
            margem = 28
            altura_fixa = (
                frame_dados.winfo_reqheight()
                + frame_feedback.winfo_reqheight()
                + botao_gerar.winfo_reqheight()
                + margem
            )
            altura_conteudo = max(300, altura_canvas - altura_fixa)
            frame_conteudo.pack_propagate(False)
            frame_conteudo.configure(height=altura_conteudo)

        if _conteudo_cabe_no_canvas():
            canvas.yview_moveto(0)

        _atualizar_scrollregion()

    # mousewheel: ativar_scroll()

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # ---------------------------- #
    # FRAME DADOS DO ORÇAMENTO     #
    # ---------------------------- #
    frame_dados = tk.LabelFrame(frame_principal, text="1. Dados do Orçamento")
    frame_dados.pack(fill="x", padx=10, pady=10)

    frame_checkboxes = tk.Frame(frame_dados)
    frame_checkboxes.grid(row=1, column=0, columnspan=2, sticky="w", padx=5)

    # linha 1 - Proprietário / Autor
    tk.Label(frame_dados, text="Autor(a):").grid(row=0, column=0, padx=5, pady=5, sticky="w")

    var_proprietario = tk.StringVar()

    def forcar_maiusculo(*args):
        texto = var_proprietario.get()
        var_proprietario.set(texto.upper())

    var_proprietario.trace_add("write", forcar_maiusculo)

    entrada_proprietario = tk.Entry(frame_dados, textvariable=var_proprietario)
    entrada_proprietario.grid(row=0, column=1, columnspan=3, sticky="ew", padx=5)

    frame_dados.columnconfigure(1, weight=1)

    tk.Label(frame_dados, text="Estado:").grid(row=0, column=4, padx=5)

    estados = ctx.obter_estados()

    combo_estado = ttk.Combobox(frame_dados, values=estados, width=6, state="readonly")
    combo_estado.grid(row=0, column=5, padx=5)

    def estado_alterado(event=None):

        if lista_anomalias:
            atualizar_tree()

    combo_estado.bind("<<ComboboxSelected>>", estado_alterado)

    def obter_estado():

        estado = combo_estado.get()

        if estado == "":
            mostrar_feedback("Selecione um Estado.", "red")
            raise ValueError

        return estado

    # linha 2 - Acompanhamento Técnico, Eventuais, Estado, Aluguel e BDI
    var_acompanhamento = tk.BooleanVar(value=True)

    chk_acompanhamento = tk.Checkbutton(
        frame_dados,
        text="Acompanhamento Técnico |",
        variable=var_acompanhamento
    )

    chk_acompanhamento.pack(in_=frame_checkboxes, side="left")

    var_eventuais = tk.BooleanVar(value=False)

    chk_eventuais = tk.Checkbutton(
        frame_dados,
        text="Eventuais (10%)",
        variable=var_eventuais
    )

    chk_eventuais.pack(in_=frame_checkboxes, side="left")

    tk.Label(frame_dados, text="Aluguel (R$):").grid(row=1, column=2, padx=5)

    entrada_aluguel = tk.Entry(frame_dados, width=10)
    entrada_aluguel.grid(row=1, column=3, padx=5)
    entrada_aluguel.insert(0, "1000")

    tk.Label(frame_dados, text="BDI (%):").grid(row=1, column=4, padx=5)

    entrada_bdi = tk.Entry(frame_dados, width=8)
    entrada_bdi.grid(row=1, column=5, padx=5)
    entrada_bdi.insert(0, "30,45")

    # ---------------------------- #
    # FRAME METRAGEM DOS CÔMODOS   #
    # ---------------------------- #
    frame_conteudo = tk.Frame(frame_principal)
    frame_conteudo.pack(fill="both", expand=True)

    frame_conteudo.columnconfigure(0, weight=0)
    frame_conteudo.columnconfigure(1, weight=1)
    frame_conteudo.rowconfigure(0, weight=1)
    frame_conteudo.rowconfigure(1, weight=1)

    frame_metragem = tk.LabelFrame(frame_conteudo, text="2. Metragem dos Cômodos")
    frame_metragem.grid(row=0, column=0, sticky="nw", padx=10, pady=5)

    frame_tabela = tk.Frame(frame_metragem)
    frame_tabela.pack(pady=5)

    tk.Label(frame_tabela, text="Cômodo", width=15, font=("Arial", 10, "bold")).grid(row=0, column=0)
    tk.Label(frame_tabela, text="Piso (m²)", width=10, font=("Arial", 10, "bold")).grid(row=0, column=1)
    tk.Label(frame_tabela, text="Rev. Arg. (m²)", width=12, font=("Arial", 10, "bold")).grid(row=0, column=2)
    tk.Label(frame_tabela, text="Rev. Cer. (m²)", width=12, font=("Arial", 10, "bold")).grid(row=0, column=3)

    lista_comodos = [
        "Sala",
        "Dormitório 1",
        "Dormitório 2",
        "Banheiro",
        "Cozinha",
        "Área de Serviço"
    ]

    comodos_area_molhada = [
        "Banheiro",
        "Cozinha",
        "Área de Serviço"
    ]

    comodos_area_seca = [
        "Sala",
        "Dormitório 1",
        "Dormitório 2"
    ]

    comodos_dr = [
        "Dormitório 1",
        "Dormitório 2",
        "Banheiro",
        "Cozinha",
        "Área de Serviço"
    ]

    comodos = {}

    for i, c in enumerate(lista_comodos, start=1):

        tk.Label(frame_tabela, text=c).grid(row=i, column=0)

        entrada_piso = tk.Entry(frame_tabela, width=10)
        entrada_piso.grid(row=i, column=1)

        entrada_rev_arg = tk.Entry(frame_tabela, width=10)
        entrada_rev_arg.grid(row=i, column=2)

        if c in comodos_area_molhada:

            entrada_rev_cer = tk.Entry(frame_tabela, width=10)

        else:

            entrada_rev_cer = tk.Entry(frame_tabela, width=10, state="disabled")

        entrada_rev_cer.grid(row=i, column=3)

        comodos[c] = {
            "piso": entrada_piso,
            "rev_arg": entrada_rev_arg,
            "rev_cer": entrada_rev_cer
        }

    # ---------------------------- #
    # FRAME SELEÇÃO DE ANOMALIA    #
    # ---------------------------- #
    frame_anomalia = tk.LabelFrame(frame_conteudo, text="3. Selecionar Anomalia")
    frame_anomalia.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

    vicios = [
        "Desplacamento de pisos cerâmicos em área seca",
        "Desplacamento de pisos cerâmicos em área molhada",
        "Desplacamento de azulejos",
        "Manchas nos pisos",
        "Manchas nos azulejos",
        "Falta de DR",
        "Infiltração pela esquadria",
        "Infiltração pela cobertura",
        "Umidade na parede",
        "Trinca saindo da janela",
        "Trinca em parede",
        "Trinca contígua à esquadria",
        "Trinca na laje percorrendo o eletroduto",
        "Pintura de parede"
    ]

    combo_vicio = ttk.Combobox(frame_anomalia, width=50, values=vicios, state="readonly")
    combo_vicio.pack(pady=5)

    comodos_bloqueados_por_vicio = {
        "Desplacamento de azulejos": comodos_area_seca,
        "Desplacamento de pisos cerâmicos em área molhada": comodos_area_seca,
        "Desplacamento de pisos cerâmicos em área seca": comodos_area_molhada,
        "Manchas nos azulejos": comodos_area_seca,
        "Manchas nos pisos": comodos_area_seca,
        "Falta de DR": comodos_dr
    }

    # ---------------------------- #
    # CHECKBOXES DE CÔMODOS        #
    # ---------------------------- #
    tk.Label(frame_anomalia, text="Cômodos afetados:").pack()

    frame_check = tk.Frame(frame_anomalia)
    frame_check.pack(pady=5)

    linhas = [
        ["Sala", "Banheiro"],
        ["Dormitório 1", "Cozinha"],
        ["Dormitório 2", "Área de Serviço"]
    ]

    checkbox_comodos = {}

    for r, linha in enumerate(linhas):

        for c, comodo in enumerate(linha):

            var = tk.BooleanVar()

            chk = tk.Checkbutton(frame_check, text=comodo, variable=var)
            chk.grid(row=r, column=c, sticky="w", padx=10)

            checkbox_comodos[comodo] = {
                "var": var,
                "widget": chk
            }

    def atualizar_checkboxes_por_vicio(event=None):

        vicio_selecionado = combo_vicio.get()

        comodos_bloqueados = comodos_bloqueados_por_vicio.get(vicio_selecionado, set())

        for comodo, dados in checkbox_comodos.items():

            var = dados["var"]
            chk = dados["widget"]

            if comodo in comodos_bloqueados:
                var.set(False)
                chk.config(state="disabled")
            else:
                chk.config(state="normal")

    combo_vicio.bind("<<ComboboxSelected>>", atualizar_checkboxes_por_vicio)

    # ---------------------------- #
    # FEEDBACK VISUAL              #
    # ---------------------------- #

    def mostrar_feedback(mensagem, cor="red", temporario=True):

        nonlocal _feedback_timer

        feedback_label.config(text=mensagem, fg=cor)

        if _feedback_timer is not None:
            root.after_cancel(_feedback_timer)

        if temporario:
            _feedback_timer = root.after(
                3000,
                lambda: feedback_label.config(text="")
        )

    # ---------------------------- #
    # LISTA DE ANOMALIAS           #
    # ---------------------------- #
    frame_lista = tk.LabelFrame(frame_conteudo, text="4. Anomalias adicionadas")
    frame_lista.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=10, pady=5)
    frame_lista.rowconfigure(0, weight=1)
    frame_lista.columnconfigure(0, weight=1)

    lista_anomalias = []

    frame_listbox = tk.Frame(frame_lista)
    frame_listbox.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)

    tree_anomalias = ttk.Treeview(
        frame_listbox,
        columns=("subtotal"),
        show="tree headings",
        height=ALTURA_TREE_MIN
    )

    tree_anomalias.heading("#0", text="Anomalia")
    tree_anomalias.heading("subtotal", text="Subtotal")

    tree_anomalias.column("#0", width=350, minwidth=200, stretch=True)
    tree_anomalias.column("subtotal", width=100, minwidth=90, stretch=False, anchor="e")

    def ajustar_altura_treeview(event=None):
        frame_listbox.update_idletasks()
        altura = frame_listbox.winfo_height()
        if altura < 40:
            return
        linhas = max(ALTURA_TREE_MIN, (altura - 28) // 20)
        tree_anomalias.configure(height=linhas)

    frame_listbox.bind("<Configure>", ajustar_altura_treeview)

    tree_anomalias.pack(side="left", fill="both", expand=True)

    scroll_lista = ttk.Scrollbar(
        frame_listbox,
        orient="vertical",
        command=tree_anomalias.yview
    )

    tree_anomalias.configure(yscrollcommand=scroll_lista.set)

    scroll_lista.pack(side="right", fill="y")

    # ---------------------------- #
    # FUNÇÃO ADICIONAR ANOMALIA    #
    # ---------------------------- #
    def adicionar_anomalia():

        vicio = combo_vicio.get()

        if not vicio:
            mostrar_feedback("Selecione uma anomalia.", "red")
            return

        comodos_afetados = [
            c for c in lista_comodos if checkbox_comodos[c]["var"].get()
        ]

        if not comodos_afetados:
            mostrar_feedback("Selecione ao menos um cômodo.", "red")
            return

        # procurar se a anomalia já foi adicionada
        for item in lista_anomalias:

            if item["vicio"] == vicio:

                novos = []

                for c in comodos_afetados:

                    if c not in item["comodos"]:
                        item["comodos"].append(c)
                        novos.append(c)

                if novos:

                    atualizar_tree()

                    mostrar_feedback(
                        f"Cômodo(s) adicionado(s): {', '.join(novos)}",
                        "orange"
                    )

                else:

                    mostrar_feedback(
                        "Esta anomalia já está cadastrada nesse(s) cômodo(s).",
                        "orange red"
                    )

                for dados in checkbox_comodos.values():
                    dados["var"].set(False)

                return

        # criar nova anomalia
        item = {
            "vicio": vicio,
            "comodos": comodos_afetados
        }

        lista_anomalias.append(item)

        atualizar_tree()

        for dados in checkbox_comodos.values():
            dados["var"].set(False)

        mostrar_feedback("Anomalia adicionada com sucesso.", "green")

    def atualizar_tree():

        tree_anomalias.delete(*tree_anomalias.get_children())

        for item in lista_anomalias:

            subtotal = calcular_subtotal_anomalia(item)

            subtotal_str = f"R$ {subtotal:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            pai = tree_anomalias.insert(
                "",
                "end",
                text=item["vicio"],
                values=(subtotal_str,),
                open=True
            )

            for comodo in item["comodos"]:

                tree_anomalias.insert(
                    pai,
                    "end",
                    text=comodo,
                    values=("",)
                )


    def aplicar_sinapi_na_interface():

        estados_disponiveis = ctx.obter_estados()

        combo_estado["values"] = estados_disponiveis

        if estados_disponiveis:
            estado_atual = combo_estado.get().strip()

        else:
            combo_estado.set("")

        ctx.atualizar_label_csv_rodape()
        ctx.atualizar_rodape()

        if lista_anomalias:
            atualizar_tree()


    # FUNÇÃO REMOVER ANOMALIA      #
    # ---------------------------- #
    def remover_anomalia():

        selecionado = tree_anomalias.selection()

        if not selecionado:
            mostrar_feedback("Selecione uma anomalia para remover.", "red")
            return

        item_id = selecionado[0]

        pai = tree_anomalias.parent(item_id)

        # SE SELECIONAR APENAS A ANOMALIA
        if pai == "":

            indice = tree_anomalias.index(item_id)
            del lista_anomalias[indice]
            atualizar_tree()
            mostrar_feedback("Anomalia removida.", "orange red")

        # SE SELECIONAR UM CÔMODO
        else:

            indice_anomalia = tree_anomalias.index(pai)
            comodo = tree_anomalias.item(item_id)["text"]
            lista_anomalias[indice_anomalia]["comodos"].remove(comodo)

            # PARA TIRAR A ANOMALIA CASO NÃO SOBRE NENHUM CÔMODO
            if not lista_anomalias[indice_anomalia]["comodos"]:
                del lista_anomalias[indice_anomalia]

            atualizar_tree()

            mostrar_feedback(f"Cômodo removido: {comodo}", "orange")

    def remover_todas_anomalias():

        if not lista_anomalias:
            mostrar_feedback("Não há anomalias para remover.", "orange")
            return

        lista_anomalias.clear()
        atualizar_tree()
        mostrar_feedback("Todas as anomalias foram removidas.", "orange red")

    # ---------------------------- #
    # BOTÕES                       #
    # ---------------------------- #
    tk.Button(
        frame_anomalia,
        text="Adicionar Anomalia",
        command=adicionar_anomalia
    ).pack(pady=10)

    tk.Button(
        frame_lista,
        text="Remover Anomalia Selecionada",
        command=remover_anomalia
    ).grid(row=1, column=0, pady=5)

    tk.Button(
        frame_lista,
        text="Remover Todas as Anomalias",
        underline=8,
        command=remover_todas_anomalias
    ).grid(row=2, column=0, pady=5)

    # ---------------------------- #
    # ÁREA DE FEEDBACK             #
    # ---------------------------- #
    frame_feedback = tk.Frame(frame_principal)
    frame_feedback.pack(pady=(5,0))

    feedback_label = tk.Label(
        frame_feedback,
        text="",
        font=("Arial", 10, "bold"),
        fg="#a67c00"
    )

    feedback_label.pack()

    # ---------------------------- #
    # FUNÇÃO CALCULAR QUANTIDADE   #
    # ---------------------------- #
    def ler_float(entry):

        valor = entry.get().strip()

        if valor == "":
            return 0

        valor = valor.replace(",", ".")

        try:
            return float(valor)
        except:
            raise ValueError

    def calcular_quantidade(etapa, medidas):

        tipo = etapa["tipo_calculo"]

        if tipo == "area_piso":
            return medidas["piso"] * etapa.get("coeficiente", 1)

        if tipo == "area_rev_arg":
            return medidas["rev_arg"] * etapa.get("coeficiente", 1)

        if tipo == "area_rev_cer":
            return medidas["rev_cer"] * etapa.get("coeficiente", 1)

        if tipo == "perimetro":
            return math.sqrt(medidas["piso"]) * 4 * etapa.get("coeficiente", 1)

        if tipo == "por_comodo":
            return etapa["coeficiente"]

        if tipo == "fixo":
            return etapa["coeficiente"]

        return 0

    def calcular_subtotal_anomalia(item):

        total = 0

        nome_anomalia = item["vicio"]
        etapas = ctx.dados_json["anomalias"][nome_anomalia]["etapas"]

        for comodo in item["comodos"]:

            piso = ler_float(comodos[comodo]["piso"])
            arg = ler_float(comodos[comodo]["rev_arg"])

            if comodos[comodo]["rev_cer"].cget("state") != "disabled":
                cer = ler_float(comodos[comodo]["rev_cer"])
            else:
                cer = 0

            medidas = {
                "piso": piso,
                "rev_arg": arg,
                "rev_cer": cer
            }

            for etapa in etapas:

                quantidade = calcular_quantidade(etapa, medidas)

                codigo = str(etapa["codigo_sinapi"])

                estado = combo_estado.get()

                linha_sinapi = ctx.sinapi[
                    (ctx.sinapi["codigo"] == codigo) &
                    (ctx.sinapi["estado"] == estado)
                ]

                if not linha_sinapi.empty:
                    valor = linha_sinapi.iloc[0]["custo"]
                else:
                    valor = 0

                total += quantidade * valor

        return total

    # ---------------- #
    # NOME DE ARQUIVOS #
    # ---------------- #
    def nome_arquivo(texto):

        texto = texto.strip()

        if not texto:
            return "sem_proprietario"

        caracteres_invalidos = '<>:"/\\|?*'

        for caractere in caracteres_invalidos:
            texto = texto.replace(caractere, "_")

        texto = "_".join(texto.split())

        return texto[:80]

    def normalizar_texto(texto):
        if texto is None:
            return ""
        texto = str(texto).strip()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(c for c in texto if not unicodedata.combining(c))
        return texto.upper()

    # ------------------- #
    # ORDEM DO ORÇAMENTO  #
    # ------------------- #
    def definir_ordem(anomalia):
        nome = normalizar_texto(anomalia)

        if "ACOMPANHAMENTO" in nome:
            return 1

        if "PISOS" in nome or "AZULEJOS" in nome:
            return 2

        if "ENTULHO" in nome or "LIMPEZA" in nome:
            return 4

        return 3

    # ---------------------------- #
    # GERAR ORÇAMENTO              #
    # ---------------------------- #
    def gerar_orcamento():

        # para diagnosticar bugs
        print("LISTA_ANOMALIAS:", lista_anomalias)

        if not lista_anomalias:
            mostrar_feedback(
                "Adicione ao menos uma anomalia para gerar o orçamento.",
                "red"
            )
            return

        linhas = []

        # controle para evitar duplicação de repintura por cômodo/código
        repintura_executada = set() # (comodo, codigo_sinapi)

        # Para ANOMALIAS selecionadas

        for item in lista_anomalias:
            if not isinstance(item, dict):
                print("Item inválido na lista:", item)
                continue

            nome_anomalia = item["vicio"]
            comodos_afetados = item["comodos"]
            dados_anomalia = ctx.dados_json["anomalias"][nome_anomalia]
            grupo_reparo = dados_anomalia.get("grupo_reparo", nome_anomalia)
            etapas = dados_anomalia["etapas"]

            for comodo in comodos_afetados:
                try:
                    piso = ler_float(comodos[comodo]["piso"])
                    arg = ler_float(comodos[comodo]["rev_arg"])

                    if comodos[comodo]["rev_cer"].cget("state") != "disabled":
                        cer = ler_float(comodos[comodo]["rev_cer"])
                    else:
                        cer = 0

                except Exception as e:
                    print("ERRO:", e)
                    mostrar_feedback(
                        f"Erro ao ler medidas do cômodo {comodo}",
                        "red"
                    )

                    return

                medidas = {
                    "piso": piso,
                    "rev_arg": arg,
                    "rev_cer": cer
                }

                for ordem, etapa in enumerate(etapas):

                    quantidade = calcular_quantidade(etapa, medidas)

                    codigo = str(etapa["codigo_sinapi"])
                    grupo_planilha = etapa.get("grupo_planilha", "")
                    # Se for repintura, só adiciona se não foi feita para este cômodo/código
                    if grupo_planilha == "repintura":
                        chave_repintura = (comodo, codigo)
                        if chave_repintura in repintura_executada:
                            continue
                        repintura_executada.add(chave_repintura)

                    estado = obter_estado()

                    linha_sinapi = ctx.sinapi[
                        (ctx.sinapi["codigo"] == codigo) &
                        (ctx.sinapi["estado"] == estado)
                    ]

                    if not linha_sinapi.empty:

                        descricao = linha_sinapi.iloc[0]["descricao"]
                        valor = linha_sinapi.iloc[0]["custo"]

                    else:

                        descricao = "Código não encontrado"
                        valor = 0

                    total = quantidade * valor

                    linhas.append({
                        "Anomalia": grupo_reparo,
                        "Grupo Planilha": grupo_planilha,
                        "Ordem": ordem,
                        "Código SINAPI": codigo,
                        "Descrição do item": descricao,
                        "Unid.": etapa["unidade"],
                        "Qtd.": round(quantidade, 2),
                        "Valor Unit.": valor,
                        "Total s/ BDI": round(total, 2)
                    })

        # Para ITENS GERAIS que irão em todos os orçamentos

        itens_gerais = ctx.dados_json.get("itens_gerais", {})

        for chave, item in itens_gerais.items():

            incluir = False

            if item.get("tipo") == "automatico":
                incluir = True

            elif item.get("tipo") == "checkbox":
                if chave == "acompanhamento_tecnico" and var_acompanhamento.get():
                    incluir = True

            if not incluir:
                continue

            for ordem, etapa in enumerate(item["etapas"]):

                quantidade = calcular_quantidade(etapa, {})

                codigo = str(etapa["codigo_sinapi"])
                estado = obter_estado()

                linha_sinapi = ctx.sinapi[
                    (ctx.sinapi["codigo"] == codigo) &
                    (ctx.sinapi["estado"] == estado)
                ]

                if not linha_sinapi.empty:
                    descricao = linha_sinapi.iloc[0]["descricao"]
                    valor = linha_sinapi.iloc[0]["custo"]
                else:
                    descricao = "Código não encontrado"
                    valor = 0

                total = quantidade * valor

                linhas.append({
                    "Anomalia": item["descricao"],
                    "Grupo Planilha": "",
                    "Ordem": ordem,
                    "Código SINAPI": codigo,
                    "Descrição do item": descricao,
                    "Unid.": etapa["unidade"],
                    "Qtd.": round(quantidade, 2),
                    "Valor Unit.": valor,
                    "Total s/ BDI": round(total, 2)
                })

        # Criação do DataFrame 
        df = pd.DataFrame(linhas)

        df = df.groupby(
            ["Anomalia", "Grupo Planilha","Código SINAPI", "Descrição do item", "Unid.", "Valor Unit."],
            as_index=False,
            sort=False
        ).agg({
            "Qtd.": "sum",
            "Total s/ BDI": "sum"
        })

        def ordem_grupo_planilha(grupo):
            if grupo == "repintura":
                return 2
            return 0

        df["Ordem_Execucao"] = df["Anomalia"].apply(definir_ordem)
        df["Ordem_Grupo"]    = df["Grupo Planilha"].apply(ordem_grupo_planilha)

        total_sem_repintura = df[df["Grupo Planilha"] != "repintura"]["Total s/ BDI"].sum()
        total_repintura     = df[df["Grupo Planilha"] == "repintura"]["Total s/ BDI"].sum()

        total_geral = total_sem_repintura + total_repintura

        df = df.sort_values(["Ordem_Execucao", "Ordem_Grupo"])

        linhas_final = []

        # itera na ordem já definida pelo sort_values
        for _, linha_df in df.iterrows():

            anomalia     = linha_df["Anomalia"]
            grupo        = linha_df["Grupo Planilha"]

            # cabeçalho da seção — só insere quando encontra a primeira linha dela
            secao_atual = ("repintura" if grupo == "repintura" else anomalia)

            if not linhas_final or linhas_final[-1].get("_secao") != secao_atual:

                if grupo == "repintura":
                    titulo_secao   = "REPINTURA APÓS INTERVENÇÕES"
                    subtotal_secao = df[df["Grupo Planilha"] == "repintura"]["Total s/ BDI"].sum()
                else:
                    titulo_secao   = NOMES_GRUPOS_REPARO.get(anomalia, anomalia).upper()
                    subtotal_secao = df[
                        (df["Anomalia"] == anomalia) &
                        (df["Grupo Planilha"] != "repintura")
                    ]["Total s/ BDI"].sum()

                linhas_final.append({
                    "_secao":            secao_atual,
                    "Código SINAPI":     "",
                    "Descrição do item": titulo_secao,
                    "Unid.": "", "Qtd.": "", "Valor Unit.": "",
                    "Total s/ BDI":      subtotal_secao
                })

            linhas_final.append({
                "_secao":            secao_atual,
                "Código SINAPI":     linha_df["Código SINAPI"],
                "Descrição do item": linha_df["Descrição do item"],
                "Unid.":             linha_df["Unid."],
                "Qtd.":              linha_df["Qtd."],
                "Valor Unit.":       linha_df["Valor Unit."],
                "Total s/ BDI":      linha_df["Total s/ BDI"]
            })

        df = pd.DataFrame(linhas_final).drop(columns=["_secao"])

        df["Total s/ BDI"] = pd.to_numeric(df["Total s/ BDI"], errors="coerce").fillna(0)

        linha_total = pd.DataFrame([{
            "Código SINAPI": "",
            "Descrição do item": "Total sem BDI",
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": round(total_geral, 2)
        }])

        df = pd.concat([df, linha_total], ignore_index=True)

        # ------------ #
        # CALCULAR BDI #
        # ------------ #
        bdi_str = entrada_bdi.get().replace(",", ".")

        try:
            bdi = float(bdi_str) / 100
        except:
            mostrar_feedback("BDI inválido.", "red")
            return

        valor_bdi = total_geral * bdi

        linha_bdi = pd.DataFrame([{
            "Código SINAPI": "",
            "Descrição do item": f"Total do BDI ({entrada_bdi.get()}%)",
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": round(valor_bdi, 2)
        }])

        df = pd.concat([df, linha_bdi], ignore_index=True)

        # --------- #
        # EVENTUAIS #
        # --------- #
        base_eventuais = total_geral + valor_bdi
        valor_eventuais = base_eventuais * 0.10 if var_eventuais.get() else 0

        if var_eventuais.get():
            linha_eventuais = pd.DataFrame([{
                "Código SINAPI": "",
                "Descrição do item": "Eventuais (10%)",
                "Unid.": "",
                "Qtd.": "",
                "Valor Unit.": "",
                "Total s/ BDI": round(valor_eventuais, 2)
            }])

            df = pd.concat([df, linha_eventuais], ignore_index=True)

        # ------- #
        # ALUGUEL #
        # ------- #
        aluguel_str = entrada_aluguel.get().replace(",", ".")

        try:
            aluguel = float(aluguel_str)
        except:
            mostrar_feedback("Valor de aluguel inválido.", "red")
            return

        linha_aluguel = pd.DataFrame([{
            "Código SINAPI": "",
            "Descrição do item": "Aluguel (1 mês)",
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": round(aluguel, 2)
        }])

        df = pd.concat([df, linha_aluguel], ignore_index=True)

        # ----------- #
        # TOTAL FINAL #
        # ----------- #
        total_final = total_geral + valor_bdi + valor_eventuais + aluguel

        linha_total_final = pd.DataFrame([{
            "Código SINAPI": "",
            "Descrição do item": "TOTAL GERAL",
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": round(total_final, 2)
        }])

        df = pd.concat([df, linha_total_final], ignore_index=True)

        df = df[["Código SINAPI", "Descrição do item", "Unid.", "Qtd.", "Valor Unit.", "Total s/ BDI"]]

        nome_proprietario = nome_arquivo(entrada_proprietario.get())
        nome_base = f"orcamento_reparos_{nome_proprietario}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

        pasta_downloads = os.path.join(os.path.expanduser("~"), "Downloads")

        if not os.path.isdir(pasta_downloads):
            pasta_downloads = os.getcwd()

        arquivo = filedialog.asksaveasfilename(
            title="Salvar orçamento como:",
            initialdir=pasta_downloads,
            initialfile=nome_base,
            defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")]
        )

        if not arquivo:
            mostrar_feedback("Geração de orçamento cancelada.", "orange", temporario=False)
            return

        try:
            df.to_excel(arquivo, index=False)
        except PermissionError:
            mostrar_feedback(
                "Feche a planilha antes de gerar novamente.",
                "red",
                temporario=False
            )
            return

        mostrar_feedback("Orçamento gerado com sucesso!", "dark green", temporario=False)

    # ---------------------------- #
    # FORMATAR PLANILHA            #
    # ---------------------------- #
        wb = load_workbook(arquivo)
        ws = wb.active

        fonte_cabecalho = Font(bold=True)

        fundo_cabecalho = PatternFill(
            start_color="006699",
            end_color="006699",
            fill_type="solid"
        )

        fundo_anomalia = PatternFill(
            start_color='D0CECE',
            end_color='D0CECE',
            fill_type='solid'
        )

        fundo_totais = PatternFill(
            start_color="F2F2F2",
            end_color="F2F2F2",
            fill_type="solid"
        )

        fundo_total_final = PatternFill(
            start_color="006699",
            end_color="006699",
            fill_type="solid"
        )

        borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6):
            for cell in row:
                cell.border = borda

        # Linha de título logo acima do cabeçalho original
        ws.insert_rows(1)
        nome_titulo = entrada_proprietario.get().strip() or ""
        ws.merge_cells("A1:F1")
        titulo_base = "ORÇAMENTO DE REPAROS DE VÍCIOS CONSTRUTIVOS"
        ws["A1"] = (
            f"{titulo_base} - {nome_titulo.upper()}" if nome_titulo else titulo_base
        )
        ws.row_dimensions[1].height = 24.75

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=5, max_col=6):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'

        # larguras das colunas
        ws.column_dimensions["A"].width = 7.24
        ws.column_dimensions["B"].width = 70
        ws.column_dimensions["C"].width = 5.96

        # largura da coluna de Qtd.
        max_length = 0

        for col in ["D"]:

            for cell in ws[col]:

                if cell.value is not None:

                    comprimento = len(str(cell.value))

                    if comprimento > max_length:
                        max_length = comprimento

        ws.column_dimensions["D"].width = max_length + 1

        # largura da coluna de Total s/ BDI
        max_length = 0

        for col in ["F"]:

            for cell in ws[col]:

                if cell.value is not None:

                    comprimento = len(str(cell.value))

                    if comprimento > max_length:
                        max_length = comprimento

        ws.column_dimensions["F"].width = max_length

        # centralizar e quebrar texto do cabeçalho
        # título (linha 1)
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = fundo_cabecalho
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # cabeçalho da tabela (linha 2)
        for cell in ws[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fundo_cabecalho
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        colunas_centro = ["A", "C", "D", "E", "F"]

        for col in colunas_centro:

            for cell in ws[col]:

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):

            for cell in row:

                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal if cell.alignment else "center",
                    vertical="center",
                    wrap_text=True
                )
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):

            codigo = row[0].value
            descricao = row[1].value

            # --------------------------- #
            # TOTAL SEM BDI, BDI, ALUGUEL #
            # --------------------------- #
            descricao_norm = normalizar_texto(descricao)

            if descricao_norm and (
                descricao_norm == "TOTAL SEM BDI" or
                descricao_norm.startswith("TOTAL DO BDI") or
                descricao_norm.startswith("EVENTUAIS") or
                descricao_norm == "ALUGUEL (1 MES)"
            ):
                linha_idx = row[0].row
                descricao_formatada = str(descricao).strip()

                if descricao_norm == "TOTAL SEM BDI":
                    descricao_formatada = "Total sem BDI"
                elif descricao_norm == "ALUGUEL (1 MES)":
                    descricao_formatada = "Aluguel (1 mês)"
                elif descricao_norm.startswith("TOTAL DO BDI"):
                    descricao_formatada = f"Total do BDI ({entrada_bdi.get()}%)"

                # O texto original está na coluna B. Após mesclar, precisa ficar em A
                ws.cell(row=linha_idx, column=1, value=descricao_formatada)
                for col_idx in range(2, 6):
                    ws.cell(row=linha_idx, column=col_idx, value=None)

                # Mesclar rótulo em A:E, como no modelo de referência.
                ws.merge_cells(start_row=linha_idx, start_column=1, end_row=linha_idx, end_column=5)

                for col_idx in range(1, 7):
                    cell = ws.cell(row=linha_idx, column=col_idx)
                    cell.fill = fundo_totais
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(
                        horizontal="right" if col_idx == 1 else "center",
                        vertical="center",
                        wrap_text=True
                    )

            # ----------- #
            # TOTAL FINAL #
            # ----------- #
            elif descricao == "TOTAL GERAL":
                linha_idx = row[0].row
                descricao_formatada = "Total do Orçamento"

                ws.cell(row=linha_idx, column=1, value=descricao_formatada)
                for col_idx in range(2, 6):
                    ws.cell(row=linha_idx, column=col_idx, value=None)

                ws.merge_cells(start_row=linha_idx, start_column=1, end_row=linha_idx, end_column=5)

                for col_idx in range(1, 7):
                    cell = ws.cell(row=linha_idx, column=col_idx)
                    cell.fill = fundo_total_final
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.alignment = Alignment(
                        horizontal="right" if col_idx == 1 else "center",
                        vertical="center",
                        wrap_text=True
                    )

            # --------------------- #
            # SUBTÍTULO DE ANOMALIA #
            # --------------------- #
            elif (codigo in ("", None)) and descricao not in ("", None):

                for cell in row:
                    cell.fill = fundo_anomalia
                    cell.font = Font(bold=True, size=12)

    # Pintar a mensagem de 'Código não encontrado' em vermelho
            descricao = str(row[1].value)

            if "Código não encontrado" in descricao:
                for cell in row:
                    cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color="FF0000"
                    )

        # ------------------ #
        # RODAPÉ DA PLANILHA #
        # ------------------ #
        linha_nota_sinapi = ws.max_row + 1
        estado_planilha = combo_estado.get().strip()
        sufixo_referencia = (
            f"{estado_planilha} {ctx.sinapi_referencia_rotulo}"
            if estado_planilha
            else ctx.sinapi_referencia_rotulo
        )
        texto_nota = f"Base de preços: SINAPI — referência: {sufixo_referencia}"
        ws.merge_cells(
            start_row=linha_nota_sinapi, start_column=1,
            end_row=linha_nota_sinapi, end_column=6,
        )
        celula_nota = ws.cell(row=linha_nota_sinapi, column=1, value=texto_nota)
        celula_nota.font = Font(size=9, italic=True, color="444444")
        celula_nota.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        wb.save(arquivo)
        os.startfile(arquivo)

    # ---------------------------- #
    # BOTÃO GERAR ORÇAMENTO        #
    # ---------------------------- #
    botao_gerar = tk.Button(
        frame_principal,
        text="Gerar Orçamento",
        font=("Arial", 11, "bold"),
        padx=12,
        pady=6,
        #   # cores do botão
        #   bg="#006699",
        #   fg="white",
        #   activebackground="#00557a",
        #   activeforeground="white",
        #   relief="flat",
        #   bd=0,
        #   cursor="hand2",

        command=gerar_orcamento
    )

    botao_gerar.pack(pady=(5,10))

    ajustar_layout_canvas._pronto = True
    canvas.bind("<Configure>", ajustar_layout_canvas)
    root.update_idletasks()
    ajustar_layout_canvas(
        type("Evt", (), {
            "width": canvas.winfo_width(),
            "height": canvas.winfo_height(),
        })()
    )
    ctx.registrar_callback_sinapi(aplicar_sinapi_na_interface)
    aplicar_sinapi_na_interface()

    def ativar_scroll():
        if not getattr(ativar_scroll, "_ativo", False):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            ativar_scroll._ativo = True

    def desativar_scroll():
        if getattr(ativar_scroll, "_ativo", False):
            root.unbind_all("<MouseWheel>")
            ativar_scroll._ativo = False

    wrapper.aplicar_sinapi = aplicar_sinapi_na_interface
    wrapper.ativar_scroll = ativar_scroll
    wrapper.desativar_scroll = desativar_scroll
    wrapper.focar = lambda: entrada_proprietario.focus()

    return wrapper
