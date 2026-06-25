from openpyxl import load_workbook
from pathlib import Path
import csv
import os

from app_paths import app_dir

# ---------------------------- #
# PASTAS                       #
# ---------------------------- #

APP_DIR = app_dir()
PASTA_REFERENCIA = APP_DIR / "sinapi" / "sinapi_referencia"
PASTA_PROCESSADO = APP_DIR / "sinapi" / "sinapi_processado"

PASTA_PROCESSADO.mkdir(parents=True, exist_ok=True)


def _nome_base_referencia(caminho_excel: Path) -> str:
    return caminho_excel.stem


def processar_arquivo(caminho_excel, callback=None):

    caminho_excel = Path(caminho_excel)
    nome_base = _nome_base_referencia(caminho_excel)

    caminho_catalogo = PASTA_PROCESSADO / f"{nome_base}_catalogo.csv"
    caminho_precos = PASTA_PROCESSADO / f"{nome_base}_precos.csv"

    def log(msg):
        print(msg)
        if callback:
            callback(msg)

    log("\n=================================")
    log(f"Processando: {caminho_excel.name}")

    if not caminho_excel.exists():
        raise FileNotFoundError(
            f"Arquivo não encontrado: {caminho_excel}"
        )

    wb = load_workbook(
        caminho_excel,
        read_only=True,
        data_only=True
    )

    catalogo: dict[str, tuple[str, str, str]] = {}
    precos: list[tuple[str, str, float]] = []

    # ========================================================== #
    # EXTRAÇÃO DAS COMPOSIÇÕES (CSD)                             #
    # ========================================================== #

    ws = wb["CSD"]

    estados = []
    colunas_custo = []

    col = 5  # coluna E

    while True:

        estado = ws.cell(row=9, column=col).value

        if not estado:
            break

        estados.append(str(estado).strip())
        colunas_custo.append(col)

        col += 2

        if col > 200:
            break

    log(f"Estados detectados ({len(estados)}): {estados}")

    log("Lendo códigos da aba Analítico...")

    ws_analitico = wb["Analítico"]

    mapa_codigos = {}

    for row in ws_analitico.iter_rows(min_row=11, values_only=True):

        tipo = row[2]
        codigo = row[1]
        descricao = row[4]
        unidade = row[5]

        if (
            tipo is None
            and isinstance(codigo, (int, float))
            and isinstance(descricao, str)
            and isinstance(unidade, str)
        ):
            chave = (descricao.strip(), unidade.strip())
            mapa_codigos[chave] = int(codigo)

    log(f"Códigos carregados: {len(mapa_codigos)} | Extraindo composições da aba CSD...")

    contador = 0
    linhas_gravadas = 0

    indices_custo = [c - 1 for c in colunas_custo]

    for row in ws.iter_rows(min_row=11, values_only=True):

        contador += 1

        descricao = row[2]
        unidade = row[3]

        if not isinstance(descricao, str):
            continue

        chave = (descricao.strip(), unidade.strip())
        codigo = mapa_codigos.get(chave)

        if not codigo:
            continue

        codigo_str = str(codigo)
        catalogo[codigo_str] = (descricao, unidade, "C")

        for estado, idx in zip(estados, indices_custo):

            custo = row[idx]

            if custo is None:
                continue

            precos.append((codigo_str, estado, float(custo)))
            linhas_gravadas += 1

        if contador % 1000 == 0:

            log(
                f"CSD: {contador} linhas analisadas | "
                f"{linhas_gravadas} preços gravados"
            )

    log("==========Extração das composições finalizada.==========")

    # ========================================================== #
    # EXTRAÇÃO DOS INSUMOS (ISD)                                 #
    # ========================================================== #

    log("\nExtraindo insumos da aba ISD...")

    ws_isd = wb["ISD"]

    linha_cabecalho = None

    for r in range(1, 50):

        valor = ws_isd.cell(row=r, column=2).value

        if valor and "Código" in str(valor):
            linha_cabecalho = r
            break

    if not linha_cabecalho:
        raise Exception("Cabeçalho da ISD não encontrado")

    log(f"Cabeçalho ISD encontrado na linha: {linha_cabecalho}")

    estados_isd = []
    colunas_custo_isd = []

    col = 6  # coluna F

    while True:

        estado = ws_isd.cell(row=linha_cabecalho, column=col).value

        if not estado:
            break

        estados_isd.append(str(estado).strip())
        colunas_custo_isd.append(col - 1)

        col += 1

        if col > 200:
            break

    log(f"Estados detectados ({len(estados_isd)}): {estados_isd}")

    linhas_isd = 0

    for row in ws_isd.iter_rows(
        min_row=linha_cabecalho + 1,
        values_only=True
    ):

        linhas_isd += 1

        codigo = row[1]
        descricao = row[2]
        unidade = row[3]

        if not codigo:
            continue

        if not isinstance(descricao, str):
            continue

        codigo_str = str(int(codigo))
        catalogo[codigo_str] = (descricao, unidade, "I")

        for estado, idx in zip(estados_isd, colunas_custo_isd):

            custo = row[idx]

            if custo is None:
                continue

            precos.append((codigo_str, estado, float(custo)))
            linhas_gravadas += 1

        if linhas_isd % 2000 == 0:

            log(
                f"ISD: {linhas_isd} linhas analisadas | "
                f"{linhas_gravadas} preços gravados"
            )

    with open(
        caminho_catalogo,
        "w",
        newline="",
        encoding="utf-8"
    ) as f_cat:
        writer = csv.writer(f_cat)
        writer.writerow(["codigo", "descricao", "unidade", "tipo"])
        for codigo, (descricao, unidade, tipo) in sorted(
            catalogo.items(), key=lambda item: int(item[0])
        ):
            writer.writerow([codigo, descricao, unidade, tipo])

    with open(
        caminho_precos,
        "w",
        newline="",
        encoding="utf-8"
    ) as f_pre:
        writer = csv.writer(f_pre)
        writer.writerow(["codigo", "estado", "custo"])
        writer.writerows(precos)

    log(f"Catálogo criado: {caminho_catalogo} ({len(catalogo)} itens)")
    log(f"Preços criados: {caminho_precos} ({len(precos)} registros)")
    log(f"Linhas CSD analisadas: {contador}")
    log(f"Linhas ISD analisadas: {linhas_isd}")
    log("\n==========Extração finalizada.==========")

    return str(caminho_catalogo)


if __name__ == "__main__":

    arquivos_excel = [
        f for f in PASTA_REFERENCIA.glob("*.xlsx")
        if not f.name.startswith("~$")
    ]

    print("Arquivos encontrados:")

    for arquivo in arquivos_excel:
        print("-", arquivo.name)

    for arquivo in arquivos_excel:

        processar_arquivo(arquivo)

    print("\n==========Script concluído.==========")
