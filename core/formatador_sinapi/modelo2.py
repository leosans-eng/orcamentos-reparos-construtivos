"""Formatação da planilha — Modelo 2 (Enviar ao Perito)."""

from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from core.formatador_sinapi.comum import (
    FORMATO_MOEDA,
    aplicar_borda_contorno,
    encontrar_linha_dados_start,
    extrair_bdi_rotulo,
    extrair_nome_obra,
    extrair_totais_finais,
    gerar_caminho_saida,
    atribuir_valor,
    planilha_ativa,
)
from core.formatador_sinapi.types import Modelo, ResultadoFormatacao


def _finalizar_total_secao(ws, row_secao, primeira_linha, ultima_linha, align_direita, linhas_totais_secoes):
    cel_secao_total = ws.cell(row=row_secao, column=6)
    cel_secao_total.alignment = align_direita
    if primeira_linha > ultima_linha:
        atribuir_valor(cel_secao_total, "A ORÇAR")
    else:
        atribuir_valor(
            cel_secao_total, f"=SUM(F{primeira_linha}:F{ultima_linha})"
        )
        cel_secao_total.number_format = FORMATO_MOEDA
        linhas_totais_secoes.append(row_secao)


def formatar_modelo2(
    caminho_origem_xlsx: str,
    *,
    caminho_saida: str | None = None,
    diretorio_saida: str | None = None,
) -> ResultadoFormatacao:
    if not os.path.isfile(caminho_origem_xlsx):
        raise FileNotFoundError(
            f"O arquivo '{caminho_origem_xlsx}' não foi encontrado."
        )

    wb_origem = openpyxl.load_workbook(caminho_origem_xlsx)
    ws_origem = planilha_ativa(wb_origem)

    nome_obra = extrair_nome_obra(ws_origem)
    bdi_rotulo = extrair_bdi_rotulo(ws_origem)
    caminho_saida_xlsx = gerar_caminho_saida(
        caminho_origem_xlsx,
        Modelo.ENVIAR_PERITO,
        nome_obra,
        caminho_saida=caminho_saida,
        diretorio_saida=diretorio_saida,
    )

    wb_destino = openpyxl.Workbook()
    ws_destino = planilha_ativa(wb_destino)
    ws_destino.title = "Orçamento Formatado"
    ws_destino.views.sheetView[0].showGridLines = True

    font_cabecalho = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    fill_cabecalho = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    font_secao = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
    fill_secao = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    font_corpo = Font(name="Aptos Narrow", size=11, bold=False, color="000000")
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_direita = Alignment(horizontal="right", vertical="center")

    top_title = "ORÇAMENTO - REPAROS DE VÍCIOS CONSTRUTIVOS - SINAPI"
    colunas_modelo = ["Código", "Descrição", "Un.", "Qtd.", "Preço Unit.", "Total sem BDI"]

    ws_destino.append([top_title] + [""] * (len(colunas_modelo) - 1))
    ws_destino.row_dimensions[1].height = 28
    ws_destino.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas_modelo))
    cel_top = ws_destino.cell(row=1, column=1)
    cel_top.font = Font(name="Aptos Narrow", size=12, bold=True, color="FFFFFF")
    cel_top.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    cel_top.alignment = Alignment(horizontal="center", vertical="center")

    ws_destino.append(colunas_modelo)
    ws_destino.row_dimensions[2].height = 22
    for col_idx, _nome_coluna in enumerate(colunas_modelo, 1):
        celula = ws_destino.cell(row=2, column=col_idx)
        celula.font = font_cabecalho
        celula.fill = fill_cabecalho
        celula.alignment = align_centro

    linha_dados_start = encontrar_linha_dados_start(ws_origem)
    ultima_secao_row = None
    primeira_linha_secao = None
    linhas_totais_secoes = []

    for row_idx in range(linha_dados_start, ws_origem.max_row + 1):
        item_val = ws_origem.cell(row=row_idx, column=1).value
        texto_extenso = str(ws_origem.cell(row=row_idx, column=4).value or "").strip()

        if (item_val is None or str(item_val).strip() == "") and texto_extenso == "":
            continue

        banco = str(ws_origem.cell(row=row_idx, column=3).value or "").strip()
        codigo_original = str(ws_origem.cell(row=row_idx, column=4).value or "").strip()
        codigo = "Comp. SINAPI" if banco.lower() == "próprio" else codigo_original
        descricao = str(ws_origem.cell(row=row_idx, column=5).value or "").strip()

        limite_linha = 85
        num_linhas = max(1, (len(descricao) // limite_linha) + 1)
        altura_linha = num_linhas * 15
        if codigo == "Comp. SINAPI":
            altura_linha = max(altura_linha, 30)

        qtd = ws_origem.cell(row=row_idx, column=7).value
        preco_s_bdi = ws_origem.cell(row=row_idx, column=8).value

        num_linha_atual = ws_destino.max_row + 1
        ws_destino.row_dimensions[num_linha_atual].height = altura_linha

        if texto_extenso and banco == "" and descricao == "":
            ws_destino.append(["", "", "", texto_extenso, "", ""])
            linha_extenso = ws_destino.max_row
            ws_destino.merge_cells(
                start_row=linha_extenso, start_column=2, end_row=linha_extenso, end_column=5
            )
            celula_extenso = ws_destino.cell(row=linha_extenso, column=2)
            celula_extenso.font = Font(name="Aptos Narrow", size=9, bold=False)
            celula_extenso.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            linhas_extenso = max(1, (len(texto_extenso) // 90) + 1)
            ws_destino.row_dimensions[linha_extenso].height = linhas_extenso * 14
            continue

        if codigo_original == "":
            if ultima_secao_row is not None and primeira_linha_secao is not None:
                _finalizar_total_secao(
                    ws_destino,
                    ultima_secao_row,
                    primeira_linha_secao,
                    num_linha_atual - 1,
                    align_direita,
                    linhas_totais_secoes,
                )

            ultima_secao_row = num_linha_atual
            primeira_linha_secao = num_linha_atual + 1
            ws_destino.append(["", descricao, "", "", "", 0])
            ws_destino.row_dimensions[num_linha_atual].height = max(altura_linha, 26)
            ws_destino.merge_cells(
                start_row=num_linha_atual, start_column=2, end_row=num_linha_atual, end_column=4
            )
            for c_idx in range(1, 7):
                celula = ws_destino.cell(row=num_linha_atual, column=c_idx)
                celula.fill = fill_secao
                celula.font = font_secao
                if c_idx == 1:
                    celula.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                elif c_idx in [2, 3]:
                    celula.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                elif c_idx == 6:
                    celula.alignment = align_direita
        else:
            un = str(ws_origem.cell(row=row_idx, column=6).value or "").strip()
            ws_destino.append(
                [codigo, descricao, un, qtd, preco_s_bdi, f"=D{num_linha_atual}*E{num_linha_atual}"]
            )
            for c_idx in range(1, 7):
                celula = ws_destino.cell(row=num_linha_atual, column=c_idx)
                celula.font = font_corpo
                if c_idx == 1:
                    celula.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                elif c_idx in [2, 3]:
                    celula.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                else:
                    celula.alignment = Alignment(horizontal="general", vertical="center")
                    if c_idx in [5, 6]:
                        celula.number_format = FORMATO_MOEDA

    if ultima_secao_row is not None and primeira_linha_secao is not None:
        if ultima_secao_row not in linhas_totais_secoes:
            _finalizar_total_secao(
                ws_destino,
                ultima_secao_row,
                primeira_linha_secao,
                ws_destino.max_row,
                align_direita,
                linhas_totais_secoes,
            )

    totais_finais = extrair_totais_finais(ws_origem)
    tabela_final = ws_destino.max_row
    if totais_finais:
        linha1 = ws_destino.max_row + 1
        ws_destino.merge_cells(start_row=linha1, start_column=1, end_row=linha1, end_column=4)
        cel = ws_destino.cell(row=linha1, column=1)
        atribuir_valor(cel, "TOTAL PARCIAL=")
        cel.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cel.font = Font(name="Aptos Narrow", size=12, bold=True)
        cel.fill = fill_secao

        ws_destino.merge_cells(start_row=linha1, start_column=5, end_row=linha1, end_column=6)
        cel_e = ws_destino.cell(row=linha1, column=5)
        cel_e.font = font_secao
        cel_e.fill = fill_secao
        cel_e.alignment = align_direita
        atribuir_valor(
            cel_e,
            (
                "=" + "+".join(f"F{row}" for row in linhas_totais_secoes)
                if linhas_totais_secoes
                else 0
            ),
        )
        cel_e.number_format = FORMATO_MOEDA

        linha2 = linha1 + 1
        ws_destino.merge_cells(start_row=linha2, start_column=1, end_row=linha2, end_column=3)
        cel = ws_destino.cell(row=linha2, column=1)
        atribuir_valor(cel, "BDI - BENEFÍCIOS E DESPESAS INDIRETAS =")
        cel.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cel.font = Font(name="Aptos Narrow", size=12, bold=True)
        cel.fill = fill_secao

        cel_d = ws_destino.cell(row=linha2, column=4)
        atribuir_valor(cel_d, f"{bdi_rotulo}%")
        cel_d.alignment = align_centro
        cel_d.font = Font(name="Aptos Narrow", size=12, bold=True)
        cel_d.fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

        ws_destino.merge_cells(start_row=linha2, start_column=5, end_row=linha2, end_column=6)
        cel_e = ws_destino.cell(row=linha2, column=5)
        cel_e.font = font_secao
        cel_e.fill = fill_secao
        cel_e.alignment = align_direita
        atribuir_valor(cel_e, f"=E{linha1}*D{linha2}")
        cel_e.number_format = FORMATO_MOEDA

        linha3 = linha2 + 1
        ws_destino.merge_cells(start_row=linha3, start_column=1, end_row=linha3, end_column=4)
        cel = ws_destino.cell(row=linha3, column=1)
        atribuir_valor(cel, "ORÇAMENTO TOTAL")
        cel.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cel.font = Font(name="Aptos Narrow", size=14, bold=True, color="FFFFFF")
        cel.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")

        ws_destino.merge_cells(start_row=linha3, start_column=5, end_row=linha3, end_column=6)
        cel_e = ws_destino.cell(row=linha3, column=5)
        cel_e.font = Font(name="Aptos Narrow", size=14, bold=True, color="FFFFFF")
        cel_e.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        cel_e.alignment = align_direita
        atribuir_valor(cel_e, f"=SUM(E{linha1}:F{linha2})")
        cel_e.number_format = FORMATO_MOEDA

        aplicar_borda_contorno(ws_destino, linha1, 1, linha3, 6)

    ws_destino.column_dimensions["A"].width = 8
    ws_destino.column_dimensions["B"].width = 80
    ws_destino.column_dimensions["C"].width = 5
    ws_destino.column_dimensions["D"].width = 8
    ws_destino.column_dimensions["E"].width = 14
    ws_destino.column_dimensions["F"].width = 18

    aplicar_borda_contorno(ws_destino, 2, 1, tabela_final, 6)
    wb_destino.save(caminho_saida_xlsx)

    return ResultadoFormatacao(
        modelo=Modelo.ENVIAR_PERITO,
        caminho_origem=os.path.abspath(caminho_origem_xlsx),
        caminho_excel=os.path.abspath(caminho_saida_xlsx),
        nome_obra=nome_obra,
    )
