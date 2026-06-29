"""Exportação de planilhas de orçamento."""

from __future__ import annotations

import os
import tempfile
import unicodedata
from datetime import datetime
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.composicoes_proprias import custo_composicao_propria_item
from core.formatador_sinapi import Modelo, formatar_planilha
from core.formatador_sinapi.comum import (
    indice_linha,
    planilha_ativa,
    sanitizar_nome_arquivo,
)
from core.orcamento_customizado import TIPO_COMPOSICAO_PROPRIA, TIPO_SINAPI
from core.planilha_sintetica import gerar_planilha_sintetica
from core.sinapi_busca import obter_item_sinapi

COLUNAS_PLANILHA = [
    "Código SINAPI",
    "Descrição do item",
    "Unid.",
    "Qtd.",
    "Valor Unit.",
    "Total s/ BDI",
]


def _sanitizar_nome_exportacao(texto: str) -> str:
    nome = sanitizar_nome_arquivo(texto or "")
    return nome if nome else "sem_nome"


def normalizar_texto(texto) -> str:
    if texto is None:
        return ""
    texto = str(texto).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.upper()


def formatar_bdi_planilha(valor: float) -> str:
    return f"{float(valor):.2f}".replace(".", ",")


def _us_para_br_numero(texto_us: str) -> str:
    return texto_us.replace(",", "X").replace(".", ",").replace("X", ".")


def _comprimento_exibicao_moeda(valor) -> int:
    """Estima largura da célula com formatação R$ #.##0,00."""
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return len(str(valor))
    texto = "R$ " + _us_para_br_numero(f"{numero:,.2f}")
    return len(texto)


def _ajustar_largura_coluna_moeda(planilha, coluna: str, largura_minima: float = 16) -> None:
    max_length = largura_minima
    for cell in planilha[coluna]:
        if cell.value is None:
            continue
        if isinstance(cell.value, (int, float)):
            max_length = max(max_length, _comprimento_exibicao_moeda(cell.value))
        else:
            max_length = max(max_length, len(str(cell.value)))
    planilha.column_dimensions[coluna].width = max_length + 2


def sincronizar_precos_sinapi(orcamento, sinapi, estado: str) -> None:
    if not estado:
        return
    for grupo in orcamento.grupos:
        for item in grupo.get("itens", []):
            if item["tipo"] != TIPO_SINAPI:
                continue
            if item.get("estado") == estado:
                continue
            linha = obter_item_sinapi(sinapi, item["codigo"], estado)
            if linha is None:
                continue
            try:
                item["custo_unitario"] = float(
                    linha.get("custo", item["custo_unitario"])
                )
            except (TypeError, ValueError):
                pass
            tipo = str(linha.get("tipo", "")).strip().upper()[:1]
            if tipo in ("I", "C"):
                item["tipo_sinapi"] = tipo
            item["estado"] = estado


def _custo_unitario_item(item, catalogo, sinapi, estado: str) -> float:
    if item["tipo"] == TIPO_SINAPI:
        return float(item["custo_unitario"])
    if item["tipo"] == TIPO_COMPOSICAO_PROPRIA:
        custo_unit, _ = custo_composicao_propria_item(
            item, catalogo, sinapi, estado
        )
        return float(custo_unit)
    return 0.0


def _descricao_item(item) -> str:
    if item["tipo"] == TIPO_SINAPI:
        return str(item.get("descricao", "")).strip()
    if item["tipo"] == TIPO_COMPOSICAO_PROPRIA:
        return str(item.get("nome", "")).strip()
    return ""


def _codigo_item(item) -> str:
    codigo = str(item.get("codigo", "")).strip()
    return codigo or "—"


def _unidade_item(item) -> str:
    return str(item.get("unidade", "")).strip()


def montar_dataframe_orcamento_customizado(orcamento, catalogo, sinapi, estado: str):
    linhas = []
    total_sem_bdi = 0.0

    for grupo in orcamento.grupos:
        subtotal_grupo = 0.0
        itens_linhas = []

        for item in grupo.get("itens", []):
            valor_unit = _custo_unitario_item(item, catalogo, sinapi, estado)
            quantidade = float(item["quantidade"])
            total_item = round(valor_unit * quantidade, 2)
            subtotal_grupo += total_item
            itens_linhas.append(
                {
                    "Código SINAPI": _codigo_item(item),
                    "Descrição do item": _descricao_item(item),
                    "Unid.": _unidade_item(item),
                    "Qtd.": round(quantidade, 2),
                    "Valor Unit.": round(valor_unit, 2),
                    "Total s/ BDI": total_item,
                }
            )

        if not grupo.get("itens"):
            continue

        subtotal_grupo = round(subtotal_grupo, 2)
        total_sem_bdi += subtotal_grupo

        linhas.append(
            {
                "Código SINAPI": "",
                "Descrição do item": grupo["nome"].strip().upper(),
                "Unid.": "",
                "Qtd.": "",
                "Valor Unit.": "",
                "Total s/ BDI": subtotal_grupo,
            }
        )
        linhas.extend(itens_linhas)

    total_sem_bdi = round(total_sem_bdi, 2)
    bdi_percent = float(orcamento.bdi_percent)
    valor_bdi = round(total_sem_bdi * bdi_percent / 100, 2)
    total_final = round(total_sem_bdi + valor_bdi, 2)

    linhas.append(
        {
            "Código SINAPI": "",
            "Descrição do item": "Total sem BDI",
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": total_sem_bdi,
        }
    )
    linhas.append(
        {
            "Código SINAPI": "",
            "Descrição do item": f"Total do BDI ({formatar_bdi_planilha(bdi_percent)}%)",
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": valor_bdi,
        }
    )
    linhas.append(
        {
            "Código SINAPI": "",
            "Descrição do item": "TOTAL GERAL",
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": total_final,
        }
    )

    df = pd.DataFrame(linhas, columns=COLUNAS_PLANILHA)
    df["Total s/ BDI"] = pd.to_numeric(df["Total s/ BDI"], errors="coerce")
    df = df.fillna({"Total s/ BDI": 0.0})
    return df


def aplicar_formato_planilha_orcamento(
    caminho_arquivo: str,
    *,
    titulo: str,
    bdi_percent: float,
    estado: str,
    referencia_sinapi: str,
) -> None:
    wb = load_workbook(caminho_arquivo)
    ws = planilha_ativa(wb)

    fundo_cabecalho = PatternFill(
        start_color="006699", end_color="006699", fill_type="solid"
    )
    fundo_anomalia = PatternFill(
        start_color="D0CECE", end_color="D0CECE", fill_type="solid"
    )
    fundo_totais = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    fundo_total_final = PatternFill(
        start_color="006699", end_color="006699", fill_type="solid"
    )
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6):
        for cell in row:
            cell.border = borda

    ws.insert_rows(1)
    ws.merge_cells("A1:F1")
    ws["A1"] = titulo
    ws.row_dimensions[1].height = 24.75

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = 'R$ #.##0,00'

    ws.column_dimensions["A"].width = 7.24
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 5.96

    for col in ["D"]:
        max_length = 0
        for cell in ws[col]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions["D"].width = max_length + 1

    for col in ["F"]:
        _ajustar_largura_coluna_moeda(ws, col)

    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = fundo_cabecalho
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fundo_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ["A", "C", "D", "E", "F"]:
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal if cell.alignment else "center",
                vertical="center",
                wrap_text=True,
            )

    bdi_texto = formatar_bdi_planilha(bdi_percent)

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        codigo = row[0].value
        descricao = row[1].value
        descricao_norm = normalizar_texto(descricao)

        if descricao_norm and (
            descricao_norm == "TOTAL SEM BDI"
            or descricao_norm.startswith("TOTAL DO BDI")
        ):
            linha_idx = indice_linha(row[0])
            descricao_formatada = str(descricao).strip()
            if descricao_norm == "TOTAL SEM BDI":
                descricao_formatada = "Total sem BDI"
            elif descricao_norm.startswith("TOTAL DO BDI"):
                descricao_formatada = f"Total do BDI ({bdi_texto}%)"

            ws.cell(row=linha_idx, column=1, value=descricao_formatada)
            for col_idx in range(2, 6):
                ws.cell(row=linha_idx, column=col_idx, value=None)
            ws.merge_cells(
                start_row=linha_idx, start_column=1, end_row=linha_idx, end_column=5
            )

            for col_idx in range(1, 7):
                cell = ws.cell(row=linha_idx, column=col_idx)
                cell.fill = fundo_totais
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(
                    horizontal="right" if col_idx == 1 else "center",
                    vertical="center",
                    wrap_text=True,
                )

        elif descricao == "TOTAL GERAL":
            linha_idx = indice_linha(row[0])
            ws.cell(row=linha_idx, column=1, value="Total do Orçamento")
            for col_idx in range(2, 6):
                ws.cell(row=linha_idx, column=col_idx, value=None)
            ws.merge_cells(
                start_row=linha_idx, start_column=1, end_row=linha_idx, end_column=5
            )

            for col_idx in range(1, 7):
                cell = ws.cell(row=linha_idx, column=col_idx)
                cell.fill = fundo_total_final
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.alignment = Alignment(
                    horizontal="right" if col_idx == 1 else "center",
                    vertical="center",
                    wrap_text=True,
                )

        elif (codigo in ("", None)) and descricao not in ("", None):
            for cell in row:
                cell.fill = fundo_anomalia
                cell.font = Font(bold=True, size=12)

        descricao_str = str(row[1].value or "")
        if "Código não encontrado" in descricao_str:
            for cell in row:
                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color="FF0000",
                )

    linha_nota_sinapi = ws.max_row + 1
    sufixo_referencia = (
        f"{estado} {referencia_sinapi}" if estado else referencia_sinapi
    )
    texto_nota = f"Base de preços: SINAPI — referência: {sufixo_referencia}"
    ws.merge_cells(
        start_row=linha_nota_sinapi,
        start_column=1,
        end_row=linha_nota_sinapi,
        end_column=6,
    )
    celula_nota = ws.cell(row=linha_nota_sinapi, column=1, value=texto_nota)
    celula_nota.font = Font(size=9, italic=True, color="444444")
    celula_nota.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb.save(caminho_arquivo)


def _validar_exportacao(parent, orcamento, estado: str) -> bool:
    if not estado:
        messagebox.showwarning(
            "Gerar planilha",
            "Selecione o estado antes de gerar a planilha.",
            parent=parent,
        )
        return False

    if not orcamento.grupos or not any(g.get("itens") for g in orcamento.grupos):
        messagebox.showwarning(
            "Gerar planilha",
            "O orçamento não possui etapas com itens para exportar.",
            parent=parent,
        )
        return False
    return True


def _pasta_downloads_padrao() -> str:
    pasta_downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    if not os.path.isdir(pasta_downloads):
        pasta_downloads = os.getcwd()
    return pasta_downloads


def _nome_obra_exportacao(orcamento) -> str:
    return (orcamento.nome or "Orçamento customizado").strip()


def exportar_orcamento_customizado_modelo_formatado(
    parent,
    modelo: int,
    orcamento,
    catalogo,
    sinapi,
    estado: str,
    referencia_sinapi: str,
) -> bool:
    if modelo not in (1, 2, 3):
        return False
    if not _validar_exportacao(parent, orcamento, estado):
        return False

    sincronizar_precos_sinapi(orcamento, sinapi, estado)
    nome_obra = _nome_obra_exportacao(orcamento)
    nome_limpo = _sanitizar_nome_exportacao(nome_obra)
    nome_base = f"Planilha_Sintetica_Convertida_Modelo{modelo} - {nome_limpo}.xlsx"

    arquivo = filedialog.asksaveasfilename(
        parent=parent,
        title="Salvar planilha como:",
        initialdir=_pasta_downloads_padrao(),
        initialfile=nome_base,
        defaultextension=".xlsx",
        filetypes=[("Planilha Excel", "*.xlsx")],
    )
    if not arquivo:
        return False

    fd, caminho_sintetica = tempfile.mkstemp(suffix=".xlsx", prefix="orc_sintetica_")
    os.close(fd)
    resultado = None
    try:
        gerar_planilha_sintetica(
            caminho_sintetica,
            orcamento,
            catalogo,
            sinapi,
            estado,
            referencia_sinapi,
        )
        resultado = formatar_planilha(
            caminho_sintetica,
            modelo=Modelo(modelo),
            caminho_saida=arquivo,
            gerar_word=modelo in (1, 3),
        )
    except PermissionError:
        messagebox.showerror(
            "Gerar planilha",
            "Feche a planilha antes de gerar novamente.",
            parent=parent,
        )
        return False
    except OSError as exc:
        messagebox.showerror(
            "Gerar planilha",
            f"Não foi possível gerar a planilha:\n{exc}",
            parent=parent,
        )
        return False
    finally:
        try:
            os.unlink(caminho_sintetica)
        except OSError:
            pass

    if resultado is None:
        return False

    try:
        os.startfile(resultado.caminho_excel)
    except OSError:
        pass

    if resultado.caminho_word:
        try:
            os.startfile(resultado.caminho_word)
        except OSError:
            pass

    mensagem = "Planilha gerada com sucesso!"
    if resultado.caminho_word:
        mensagem += "\n\nDocumento Word gerado no mesmo local."
    if resultado.avisos:
        mensagem += "\n\n" + "\n".join(resultado.avisos)

    messagebox.showinfo("Gerar planilha", mensagem, parent=parent)
    return True


def exportar_orcamento_customizado_modelo4(
    parent,
    orcamento,
    catalogo,
    sinapi,
    estado: str,
    referencia_sinapi: str,
) -> bool:
    if not _validar_exportacao(parent, orcamento, estado):
        return False

    sincronizar_precos_sinapi(orcamento, sinapi, estado)
    df = montar_dataframe_orcamento_customizado(orcamento, catalogo, sinapi, estado)

    nome_base = (
        f"orcamento_customizado_{_sanitizar_nome_exportacao(orcamento.nome)}_"
        f"{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    )

    arquivo = filedialog.asksaveasfilename(
        parent=parent,
        title="Salvar planilha como:",
        initialdir=_pasta_downloads_padrao(),
        initialfile=nome_base,
        defaultextension=".xlsx",
        filetypes=[("Planilha Excel", "*.xlsx")],
    )
    if not arquivo:
        return False

    try:
        df.to_excel(arquivo, index=False)
    except PermissionError:
        messagebox.showerror(
            "Gerar planilha",
            "Feche a planilha antes de gerar novamente.",
            parent=parent,
        )
        return False

    titulo_base = "ORÇAMENTO CUSTOMIZADO"
    nome_orcamento = (orcamento.nome or "").strip()
    titulo = (
        f"{titulo_base} - {nome_orcamento.upper()}" if nome_orcamento else titulo_base
    )

    try:
        aplicar_formato_planilha_orcamento(
            arquivo,
            titulo=titulo,
            bdi_percent=float(orcamento.bdi_percent),
            estado=estado,
            referencia_sinapi=referencia_sinapi,
        )
    except OSError as exc:
        messagebox.showerror(
            "Gerar planilha",
            f"Não foi possível formatar a planilha:\n{exc}",
            parent=parent,
        )
        return False

    try:
        os.startfile(arquivo)
    except OSError:
        pass

    messagebox.showinfo(
        "Gerar planilha",
        "Planilha gerada com sucesso!",
        parent=parent,
    )
    return True
