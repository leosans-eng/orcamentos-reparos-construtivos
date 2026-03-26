import tkinter as tk
from tkinter import ttk, filedialog
import pandas as pd
import json
import math
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import os
import unicodedata
from datetime import datetime

# ------------------------------------------- #
# VERSÃO DO SISTEMA (INTERFACE E EXPORTAÇÕES) #
# ------------------------------------------- #

APP_VERSION = "0.9.5.1"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PASTA_SINAPI_PROCESSADO = os.path.join(BASE_DIR, "sinapi", "sinapi_processado")
CAMINHO_FALLBACK_SINAPI = os.path.join(BASE_DIR, "sinapi_precos.csv")

def _parse_referencia_do_nome_csv(nome_arquivo):
    #Extrai (ano, mês) de SINAPI_Referência_YYYY_MM.csv (acento opcional em Referência)
    m = re.match(
        r"(?i)SINAPI_Refer[eê]ncia_(\d{4})_(\d{2})\.csv$",
        nome_arquivo.strip(),
    )
    if not m:
        return None
    ano, mes = int(m.group(1)), int(m.group(2))
    if not (1 <= mes <= 12):
        return None
    return ano, mes

def obter_csv_sinapi_mais_recente(pasta_processado):
    """
    Retorna (caminho_absoluto, rotulo_referencia) do CSV mais recente na pasta processada.
    rotulo_referencia: texto para interface/planilha, ex. '03/2026'.
    """
    if not os.path.isdir(pasta_processado):
        return None, None
    candidatos = []
    for nome in os.listdir(pasta_processado):
        if not nome.lower().endswith(".csv"):
            continue
        tupla_data = _parse_referencia_do_nome_csv(nome)
        if tupla_data:
            candidatos.append(
                (tupla_data, os.path.join(pasta_processado, nome), nome)
            )
    if not candidatos:
        return None, None
    candidatos.sort(key=lambda x: x[0], reverse=True)
    (ano, mes), caminho, _ = candidatos[0]
    rotulo = f"{mes:02d}/{ano}"
    return caminho, rotulo

def informacoes_versao():
    return {
        "app": APP_VERSION,
        "sinapi": sinapi_referencia_rotulo,
        "arquivo_sinapi": os.path.basename(caminho_sinapi_carregado),
    }

def texto_rodape_interface():
    info = informacoes_versao()
    return f"Sistema ORC v{info['app']} · SINAPI referência {info['sinapi']}"

# Atraso antes do nome do CSV sair deslizando à direita (ms)
RODAPE_CSV_SUMIR_APOS_MS = 3000
# Intervalo e passo do deslize (quanto menor o intervalo, mais suave)
RODAPE_CSV_DESLIZE_INTERVALO_MS = 35
RODAPE_CSV_DESLIZE_PASSO_PX = 4

def _agendar_sumico_nome_csv_deslize_direita(label_csv, frame_rodape):

    def executar():
        if not label_csv.winfo_exists():
            return
        if label_csv.winfo_manager() != "pack":
            return
        label_csv.pack_forget()
        label_csv.place(relx=1.0, rely=0.5, anchor="e", x=-6, in_=frame_rodape)

        def deslizar():
            if not label_csv.winfo_exists():
                return
            frame_rodape.update_idletasks()
            label_csv.update_idletasks()
            limite = frame_rodape.winfo_width()
            # Posição à esquerda do widget em relação ao rodapé (sai totalmente à direita)
            if label_csv.winfo_x() - frame_rodape.winfo_x() > limite + 8:
                label_csv.destroy()
                return
            info = label_csv.place_info()
            cur_x = int(float(info.get("x", 0)))
            label_csv.place(
                relx=1.0,
                rely=0.5,
                anchor="e",
                x=cur_x + RODAPE_CSV_DESLIZE_PASSO_PX,
                in_=frame_rodape,
            )
            frame_rodape.after(RODAPE_CSV_DESLIZE_INTERVALO_MS, deslizar)

        frame_rodape.after(25, deslizar)

    frame_rodape.after(RODAPE_CSV_SUMIR_APOS_MS, executar)

# ---------------------------- #
# CARREGAR DADOS JSON e CSV    #
# ---------------------------- #

with open("vicios_construtivos.json", "r", encoding="utf-8") as f:
    dados_json = json.load(f)

if isinstance(dados_json, list):
    dados_json = dados_json[0]

nomes_grupos_reparo = {
    "reparo_pisos_ceramicos": "Reparo de Pisos Cerâmicos",
    "reparo_azulejos": "Reparo de Azulejos",
    "reparo_trincas": "Intervenção de fissuras e trincas"
}

caminho_sinapi_carregado, sinapi_referencia_rotulo = obter_csv_sinapi_mais_recente(
    PASTA_SINAPI_PROCESSADO
)

if caminho_sinapi_carregado is None and os.path.isfile(CAMINHO_FALLBACK_SINAPI):
    caminho_sinapi_carregado = CAMINHO_FALLBACK_SINAPI
    sinapi_referencia_rotulo = "arquivo local (sinapi_precos.csv na raiz do projeto)"

if caminho_sinapi_carregado is None:
    raise SystemExit(
        "Não foi encontrado CSV SINAPI em sinapi/sinapi_processado "
        "(padrão: SINAPI_Referência_AAAA_MM.csv) e o arquivo de contingência "
        f"sinapi_precos.csv não existe em:\n{BASE_DIR}"
    )

sinapi = pd.read_csv(caminho_sinapi_carregado, dtype={"codigo": str})

sinapi.columns = sinapi.columns.str.strip().str.lower()

# ---------------------------- #
# JANELA PRINCIPAL             #
# ---------------------------- #

janela = tk.Tk()
janela.title("Orçamento de Reparos Construtivos - ORC")
janela.geometry("990x610+200+40")

# ---------------------------- #
# RODAPÉ (VERSÕES / SINAPI)    #
# ---------------------------- #

frame_rodape = tk.Frame(janela)
frame_rodape.pack(side="bottom", fill="x", padx=10, pady=(0, 6))

tk.Label(
    frame_rodape,
    text=texto_rodape_interface(),
    font=("Arial", 8),
    fg="#555555",
    anchor="w",
).pack(side="left", anchor="w")

label_nome_csv_rodape = tk.Label(
    frame_rodape,
    text=f"{os.path.basename(caminho_sinapi_carregado)} ⭠ Arquivo de base",
    font=("Arial", 8, "bold"),
    fg="#C62828",
    anchor="e",
)
label_nome_csv_rodape.pack(side="right", anchor="e")

_agendar_sumico_nome_csv_deslize_direita(label_nome_csv_rodape, frame_rodape)

# ---------------------------- #
# SCROLL DA JANELA             #
# ---------------------------- #

container = tk.Frame(janela)
container.pack(fill="both", expand=True)

canvas = tk.Canvas(container)
scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)

frame_principal = tk.Frame(canvas)

frame_principal.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas_window = canvas.create_window((0, 0), window=frame_principal, anchor="nw")

canvas.configure(yscrollcommand=scrollbar.set)

def atualizar_scroll(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

def _on_mousewheel(event):

    widget = janela.winfo_containing(event.x_root, event.y_root)

    if isinstance(widget, ttk.Combobox):
        return

    pos = canvas.yview()

    if event.delta > 0 and pos[0] <= 0:
        return

    canvas.yview_scroll(int(-1*(event.delta/120)), "units")

def ajustar_largura(event):
    canvas.itemconfig(canvas_window, width=event.width)

canvas.bind("<Configure>", ajustar_largura)

canvas.bind_all("<MouseWheel>", _on_mousewheel)

def ajustar_canvas(event):
    canvas.itemconfig(canvas_window, width=event.width)

canvas.bind("<Configure>", ajustar_canvas)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# ---------------------------- #
# FRAME DADOS DO ORÇAMENTO     #
# ---------------------------- #

frame_dados = tk.LabelFrame(frame_principal, text="Dados do Orçamento")
frame_dados.pack(fill="x", padx=10, pady=10)

frame_checkboxes = tk.Frame(frame_dados)
frame_checkboxes.grid(row=1, column=0, columnspan=2, sticky="w", padx=5)

# linha 1 - Proprietário / Autor
tk.Label(frame_dados, text="Autor(a):").grid(row=0, column=0, padx=5, pady=5, sticky="w")

entrada_proprietario = tk.Entry(frame_dados)
entrada_proprietario.grid(row=0, column=1, columnspan=3, sticky="ew", padx=5)

frame_dados.columnconfigure(1, weight=1)

tk.Label(frame_dados, text="Estado:").grid(row=0, column=4, padx=5)

estados = sorted(sinapi["estado"].dropna().unique())

combo_estado = ttk.Combobox(frame_dados, values=estados, width=6, state="readonly")
combo_estado.grid(row=0, column=5, padx=5)

def estado_alterado(event):

    if lista_anomalias:
        atualizar_tree()

combo_estado.bind("<<ComboboxSelected>>", estado_alterado)

def obter_estado():

    estado = combo_estado.get()

    if estado == "":
        mostrar_feedback("Selecione um Estado.", "red")
        raise ValueError

    return estado

# linha 2 - Acompanhamento Técnico, Eventuais, Estado, Aluguel e BDI

var_acompanhamento = tk.BooleanVar(value=True)

chk_acompanhamento = tk.Checkbutton(
    frame_dados,
    text="Acompanhamento Técnico |",
    variable=var_acompanhamento
)

chk_acompanhamento.pack(in_=frame_checkboxes, side="left")

var_eventuais = tk.BooleanVar(value=True)

chk_eventuais = tk.Checkbutton(
    frame_dados,
    text="Eventuais (10%)",
    variable=var_eventuais
)

chk_eventuais.pack(in_=frame_checkboxes, side="left")

tk.Label(frame_dados, text="Aluguel (R$):").grid(row=1, column=2, padx=5)

entrada_aluguel = tk.Entry(frame_dados, width=10)
entrada_aluguel.grid(row=1, column=3, padx=5)
entrada_aluguel.insert(0, "1000")

tk.Label(frame_dados, text="BDI (%):").grid(row=1, column=4, padx=5)

entrada_bdi = tk.Entry(frame_dados, width=8)
entrada_bdi.grid(row=1, column=5, padx=5)
entrada_bdi.insert(0, "30,45")

# ---------------------------- #
# FRAME METRAGEM DOS CÔMODOS   #
# ---------------------------- #

frame_conteudo = tk.Frame(frame_principal)
frame_conteudo.pack(fill="both", expand=True)

frame_conteudo.columnconfigure(0, weight=0)
frame_conteudo.columnconfigure(1, weight=1)

frame_metragem = tk.LabelFrame(frame_conteudo, text="Metragem dos Cômodos")
frame_metragem.grid(row=0, column=0, sticky="nw", padx=10, pady=5)

frame_tabela = tk.Frame(frame_metragem)
frame_tabela.pack(pady=5)

tk.Label(frame_tabela, text="Cômodo", width=15, font=("Arial", 10, "bold")).grid(row=0, column=0)
tk.Label(frame_tabela, text="Piso (m²)", width=10, font=("Arial", 10, "bold")).grid(row=0, column=1)
tk.Label(frame_tabela, text="Rev. Arg. (m²)", width=12, font=("Arial", 10, "bold")).grid(row=0, column=2)
tk.Label(frame_tabela, text="Rev. Cer. (m²)", width=12, font=("Arial", 10, "bold")).grid(row=0, column=3)

lista_comodos = [
    "Sala",
    "Dormitório 1",
    "Dormitório 2",
    "Banheiro",
    "Cozinha",
    "Área de Serviço"
]

comodos_area_molhada = [
    "Banheiro",
    "Cozinha",
    "Área de Serviço"
]

comodos_area_seca = [
    "Sala",
    "Dormitório 1",
    "Dormitório 2"
]

comodos = {}

for i, c in enumerate(lista_comodos, start=1):

    tk.Label(frame_tabela, text=c).grid(row=i, column=0)

    entrada_piso = tk.Entry(frame_tabela, width=10)
    entrada_piso.grid(row=i, column=1)

    entrada_rev_arg = tk.Entry(frame_tabela, width=10)
    entrada_rev_arg.grid(row=i, column=2)

    if c in comodos_area_molhada:

        entrada_rev_cer = tk.Entry(frame_tabela, width=10)

    else:

        entrada_rev_cer = tk.Entry(frame_tabela, width=10, state="disabled")

    entrada_rev_cer.grid(row=i, column=3)

    comodos[c] = {
        "piso": entrada_piso,
        "rev_arg": entrada_rev_arg,
        "rev_cer": entrada_rev_cer
    }

# ---------------------------- #
# FRAME SELEÇÃO DE ANOMALIA    #
# ---------------------------- #

frame_anomalia = tk.LabelFrame(frame_conteudo, text="Selecionar Anomalia")
frame_anomalia.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

vicios = [
    "Desplacamento de pisos cerâmicos",
    "Desplacamento de pisos cerâmicos em área molhada",
    "Desplacamento de azulejos",
    "Manchas nos pisos",
    "Manchas nos azulejos",
    "Infiltração pela esquadria",
    "Trinca na laje percorrendo o eletroduto"
]

combo_vicio = ttk.Combobox(frame_anomalia, width=40, values=vicios, state="readonly")
combo_vicio.pack(pady=5)

comodos_bloqueados_por_vicio = {
    "Desplacamento de azulejos": comodos_area_seca,
    "Desplacamento de pisos cerâmicos em área molhada": comodos_area_seca,
    "Desplacamento de pisos cerâmicos": comodos_area_molhada,
    "Manchas nos azulejos": comodos_area_seca,
    "Manchas nos pisos": comodos_area_seca
}

# ---------------------------- #
# CHECKBOXES DE CÔMODOS        #
# ---------------------------- #

tk.Label(frame_anomalia, text="Cômodos afetados:").pack()

frame_check = tk.Frame(frame_anomalia)
frame_check.pack(pady=5)

linhas = [
    ["Sala", "Banheiro"],
    ["Dormitório 1", "Cozinha"],
    ["Dormitório 2", "Área de Serviço"]
]

checkbox_comodos = {}

for r, linha in enumerate(linhas):

    for c, comodo in enumerate(linha):

        var = tk.BooleanVar()

        chk = tk.Checkbutton(frame_check, text=comodo, variable=var)
        chk.grid(row=r, column=c, sticky="w", padx=10)

        checkbox_comodos[comodo] = {
            "var": var,
            "widget": chk
        }

def atualizar_checkboxes_por_vicio(event=None):

    vicio_selecionado = combo_vicio.get()

    comodos_bloqueados = comodos_bloqueados_por_vicio.get(vicio_selecionado, set())

    for comodo, dados in checkbox_comodos.items():

        var = dados["var"]
        chk = dados["widget"]

        if comodo in comodos_bloqueados:
            var.set(False)
            chk.config(state="disabled")
        else:
            chk.config(state="normal")

combo_vicio.bind("<<ComboboxSelected>>", atualizar_checkboxes_por_vicio)

# ---------------------------- #
# FEEDBACK VISUAL              #
# ---------------------------- #

feedback_timer = None

def mostrar_feedback(mensagem, cor="red", temporario=True):

    global feedback_timer

    feedback_label.config(text=mensagem, fg=cor)

    if feedback_timer is not None:
        janela.after_cancel(feedback_timer)

    if temporario:
        feedback_timer = janela.after(
            3000,
            lambda: feedback_label.config(text="")
    )

# ---------------------------- #
# LISTA DE ANOMALIAS           #
# ---------------------------- #

frame_lista = tk.LabelFrame(frame_conteudo, text="Anomalias adicionadas")
frame_lista.grid(row=0, column=1, rowspan=2, sticky="n", padx=10, pady=5)

lista_anomalias = []

frame_listbox = tk.Frame(frame_lista)
frame_listbox.pack(fill="both", expand=True)

tree_anomalias = ttk.Treeview(
    frame_listbox,
    columns=("subtotal"),
    show="tree headings",
    height=11
)

tree_anomalias.heading("#0", text="Anomalia")
tree_anomalias.heading("subtotal", text="Subtotal")

tree_anomalias.column("#0", width=350)
tree_anomalias.column("subtotal", width=100, anchor="e")

tree_anomalias.pack(side="left", fill="both", expand=True)

scroll_lista = ttk.Scrollbar(
    frame_listbox,
    orient="vertical",
    command=tree_anomalias.yview
)

tree_anomalias.configure(yscrollcommand=scroll_lista.set)

scroll_lista.pack(side="right", fill="y")

# ---------------------------- #
# FUNÇÃO ADICIONAR ANOMALIA    #
# ---------------------------- #

def adicionar_anomalia():

    vicio = combo_vicio.get()

    if not vicio:
        mostrar_feedback("Selecione uma anomalia.", "red")
        return

    comodos_afetados = [
        c for c in lista_comodos if checkbox_comodos[c]["var"].get()
    ]

    if not comodos_afetados:
        mostrar_feedback("Selecione ao menos um cômodo.", "red")
        return
    
    # procurar se a anomalia já foi adicionada
    for item in lista_anomalias:

        if item["vicio"] == vicio:

            novos = []

            for c in comodos_afetados:

                if c not in item["comodos"]:
                    item["comodos"].append(c)
                    novos.append(c)

            if novos:

                atualizar_tree()

                mostrar_feedback(
                    f"Cômodo(s) adicionado(s): {', '.join(novos)}",
                    "orange"
                )

            else:

                mostrar_feedback(
                    "Esta anomalia já está cadastrada nesse(s) cômodo(s).",
                    "orange red"
                )

            for dados in checkbox_comodos.values():
                dados["var"].set(False)

            return
        
    # criar nova anomalia
    item = {
        "vicio": vicio,
        "comodos": comodos_afetados
    }

    lista_anomalias.append(item)

    atualizar_tree()

    for dados in checkbox_comodos.values():
        dados["var"].set(False)

    mostrar_feedback("Anomalia adicionada com sucesso.", "green")

def atualizar_tree():

    tree_anomalias.delete(*tree_anomalias.get_children())

    for item in lista_anomalias:

        subtotal = calcular_subtotal_anomalia(item)

        subtotal_str = f"R$ {subtotal:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        pai = tree_anomalias.insert(
            "",
            "end",
            text=item["vicio"],
            values=(subtotal_str,),
            open=True
        )

        for comodo in item["comodos"]:

            tree_anomalias.insert(
                pai,
                "end",
                text=comodo,
                values=("",)
            )
# ---------------------------- #
# FUNÇÃO REMOVER ANOMALIA      #
# ---------------------------- #

def remover_anomalia():

    selecionado = tree_anomalias.selection()

    if not selecionado:
        mostrar_feedback("Selecione uma anomalia para remover.", "red")
        return

    item_id = selecionado[0]

    pai = tree_anomalias.parent(item_id)

    # SE SELECIONAR APENAS A ANOMALIA
    if pai == "":

        indice = tree_anomalias.index(item_id)
        del lista_anomalias[indice]
        atualizar_tree()
        mostrar_feedback("Anomalia removida.", "orange red")

    # SE SELECIONAR UM CÔMODO
    else:

        indice_anomalia = tree_anomalias.index(pai)
        comodo = tree_anomalias.item(item_id)["text"]
        lista_anomalias[indice_anomalia]["comodos"].remove(comodo)

        # PARA TIRAR A ANOMALIA CASO NÃO SOBRE NENHUM CÔMODO
        if not lista_anomalias[indice_anomalia]["comodos"]:
            del lista_anomalias[indice_anomalia]

        atualizar_tree()

        mostrar_feedback(f"Cômodo removido: {comodo}", "orange")

def remover_todas_anomalias():

    if not lista_anomalias:
        mostrar_feedback("Não há anomalias para remover.", "orange")
        return

    lista_anomalias.clear()
    atualizar_tree()
    mostrar_feedback("Todas as anomalias foram removidas.", "orange red")
    
# ---------------------------- #
# BOTÕES                       #
# ---------------------------- #

tk.Button(
    frame_anomalia,
    text="Adicionar Anomalia",
    command=adicionar_anomalia
).pack(pady=10)

tk.Button(
    frame_lista,
    text="Remover Anomalia Selecionada",
    command=remover_anomalia
).pack(pady=5)

tk.Button(
    frame_lista,
    text="Remover Todas as Anomalias",
    underline=8,
    command=remover_todas_anomalias
).pack(pady=5)

# ---------------------------- #
# ÁREA DE FEEDBACK             #
# ---------------------------- #

frame_feedback = tk.Frame(frame_principal)
frame_feedback.pack(pady=(5,0))

feedback_label = tk.Label(
    frame_feedback,
    text="",
    font=("Arial", 10, "bold"),
    fg="#a67c00"
)

feedback_label.pack()

# ---------------------------- #
# FUNÇÃO CALCULAR QUANTIDADE   #
# ---------------------------- #

def ler_float(entry):

    valor = entry.get().strip()

    if valor == "":
        return 0

    valor = valor.replace(",", ".")

    try:
        return float(valor)
    except:
        raise ValueError
    
def calcular_quantidade(etapa, medidas):

    tipo = etapa["tipo_calculo"]

    if tipo == "area_piso":
        return medidas["piso"] * etapa.get("coeficiente", 1)

    if tipo == "area_rev_arg":
        return medidas["rev_arg"] * etapa.get("coeficiente", 1)

    if tipo == "area_rev_cer":
        return medidas["rev_cer"] * etapa.get("coeficiente", 1)

    if tipo == "perimetro":
        return math.sqrt(medidas["piso"]) * 4 * etapa.get("coeficiente", 1)

    if tipo == "por_comodo":
        return etapa["coeficiente"]

    if tipo == "fixo":
        return etapa["coeficiente"]

    return 0

def calcular_subtotal_anomalia(item):

    total = 0

    nome_anomalia = item["vicio"]
    etapas = dados_json["anomalias"][nome_anomalia]["etapas"]

    for comodo in item["comodos"]:

        piso = ler_float(comodos[comodo]["piso"])
        arg = ler_float(comodos[comodo]["rev_arg"])

        if comodos[comodo]["rev_cer"].cget("state") != "disabled":
            cer = ler_float(comodos[comodo]["rev_cer"])
        else:
            cer = 0

        medidas = {
            "piso": piso,
            "rev_arg": arg,
            "rev_cer": cer
        }

        for etapa in etapas:

            quantidade = calcular_quantidade(etapa, medidas)

            codigo = str(etapa["codigo_sinapi"])

            estado = combo_estado.get()

            linha_sinapi = sinapi[
                (sinapi["codigo"] == codigo) &
                (sinapi["estado"] == estado)
            ]

            if not linha_sinapi.empty:
                valor = linha_sinapi.iloc[0]["custo"]
            else:
                valor = 0

            total += quantidade * valor

    return total

# ---------------- #
# NOME DE ARQUIVOS #
# ---------------- #

def nome_arquivo(texto):

    texto = texto.strip()

    if not texto:
        return "sem_proprietario"

    caracteres_invalidos = '<>:"/\\|?*'

    for caractere in caracteres_invalidos:
        texto = texto.replace(caractere, "_")

    texto = "_".join(texto.split())

    return texto[:80]

def normalizar_texto(texto):
    if texto is None:
        return ""
    texto = str(texto).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.upper()

# ------------------- #
# ORDEM DO ORÇAMENTO  #
# ------------------- #

def definir_ordem(anomalia):
    nome = normalizar_texto(anomalia)

    if "ACOMPANHAMENTO" in nome:
        return 1

    if "ENTULHO" in nome or "LIMPEZA" in nome:
        return 3

    return 2

# ---------------------------- #
# GERAR ORÇAMENTO              #
# ---------------------------- #

def gerar_orcamento():

    # para diagnosticar bugs
    print("LISTA_ANOMALIAS:", lista_anomalias)

    if not lista_anomalias:
        mostrar_feedback(
            "Adicione ao menos uma anomalia para gerar o orçamento.",
            "red"
        )
        return

    linhas = []

    # controle para evitar duplicação de reparos
    reparos_executados = set()

    # Para ANOMALIAS selecionadas

    for item in lista_anomalias:

        if not isinstance(item, dict):
            print("Item inválido na lista:", item)
            continue

        nome_anomalia = item["vicio"]
        comodos_afetados = item["comodos"]

        print("Anomalia selecionada:", nome_anomalia)
        print("Chaves disponíveis no JSON:", dados_json["anomalias"].keys())

        dados_anomalia = dados_json["anomalias"][nome_anomalia]

        grupo_reparo = dados_anomalia.get("grupo_reparo", nome_anomalia)
        print("GRUPO DE REPARO:", grupo_reparo)

        etapas = dados_anomalia["etapas"]

        for comodo in comodos_afetados:
            
            chave_reparo = (grupo_reparo, comodo)
            
            if chave_reparo in reparos_executados:
                continue

            reparos_executados.add(chave_reparo)

            try:

                piso = ler_float(comodos[comodo]["piso"])
                arg = ler_float(comodos[comodo]["rev_arg"])

                if comodos[comodo]["rev_cer"].cget("state") != "disabled":
                    cer = ler_float(comodos[comodo]["rev_cer"])
                else:
                    cer = 0

            except Exception as e:

                print("ERRO:", e)

                mostrar_feedback(
                    f"Erro ao ler medidas do cômodo {comodo}",
                    "red"
                )

                return

            medidas = {
                "piso": piso,
                "rev_arg": arg,
                "rev_cer": cer
            }

            for ordem, etapa in enumerate(etapas):

                quantidade = calcular_quantidade(etapa, medidas)

                codigo = str(etapa["codigo_sinapi"])

                estado = obter_estado()

                linha_sinapi = sinapi[
                    (sinapi["codigo"] == codigo) &
                    (sinapi["estado"] == estado)
                ]

                if not linha_sinapi.empty:

                    descricao = linha_sinapi.iloc[0]["descricao"]
                    valor = linha_sinapi.iloc[0]["custo"]

                else:

                    descricao = "Código não encontrado"
                    valor = 0

                total = quantidade * valor

                linhas.append({
                    "Anomalia": grupo_reparo,
                    "Ordem": ordem,
                    "Código SINAPI": codigo,
                    "Descrição do item": descricao,
                    "Unid.": etapa["unidade"],
                    "Qtd.": round(quantidade, 2),
                    "Valor Unit.": valor,
                    "Total s/ BDI": round(total, 2)
                })

    # Para ITENS GERAIS que irão em todos os orçamentos
    
    itens_gerais = dados_json.get("itens_gerais", {})

    for chave, item in itens_gerais.items():

        incluir = False

        if item.get("tipo") == "automatico":
            incluir = True

        elif item.get("tipo") == "checkbox":
            if chave == "acompanhamento_tecnico" and var_acompanhamento.get():
                incluir = True

        if not incluir:
            continue

        for ordem, etapa in enumerate(item["etapas"]):

            quantidade = calcular_quantidade(etapa, {})

            codigo = str(etapa["codigo_sinapi"])
            estado = obter_estado()

            linha_sinapi = sinapi[
                (sinapi["codigo"] == codigo) &
                (sinapi["estado"] == estado)
            ]

            if not linha_sinapi.empty:
                descricao = linha_sinapi.iloc[0]["descricao"]
                valor = linha_sinapi.iloc[0]["custo"]
            else:
                descricao = "Código não encontrado"
                valor = 0

            total = quantidade * valor

            linhas.append({
                "Anomalia": item["descricao"],
                "Ordem": ordem,
                "Código SINAPI": codigo,
                "Descrição do item": descricao,
                "Unid.": etapa["unidade"],
                "Qtd.": round(quantidade, 2),
                "Valor Unit.": valor,
                "Total s/ BDI": round(total, 2)
            })

    # Criação do DataFrame 
    df = pd.DataFrame(linhas)

    df = df.groupby(
        ["Anomalia", "Código SINAPI", "Descrição do item", "Unid.", "Valor Unit."],
        as_index=False,
        sort=False
    ).agg({
        "Qtd.": "sum",
        "Total s/ BDI": "sum"
    })

    df["Ordem_Execucao"] = df["Anomalia"].apply(definir_ordem)

    total_geral = df["Total s/ BDI"].sum()

    df = df.sort_values(["Ordem_Execucao"])

    linhas_final = []

    for anomalia, grupo in df.groupby("Anomalia", sort=False):

        subtotal = grupo["Total s/ BDI"].sum()

        titulo = nomes_grupos_reparo.get(anomalia, anomalia)

        # subtítulo da anomalia
        linhas_final.append({
            "Código SINAPI": "",
            "Descrição do item": titulo.upper(),
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": subtotal
        })

        # itens SINAPI
        for _, row in grupo.iterrows():

            linhas_final.append({
                "Código SINAPI": row["Código SINAPI"],
                "Descrição do item": row["Descrição do item"],
                "Unid.": row["Unid."],
                "Qtd.": row["Qtd."],
                "Valor Unit.": row["Valor Unit."],
                "Total s/ BDI": row["Total s/ BDI"]
            })

    df = pd.DataFrame(linhas_final)

    df["Total s/ BDI"] = pd.to_numeric(df["Total s/ BDI"], errors="coerce").fillna(0)

    linha_total = pd.DataFrame([{
        "Código SINAPI": "",
        "Descrição do item": "Total sem BDI",
        "Unid.": "",
        "Qtd.": "",
        "Valor Unit.": "",
        "Total s/ BDI": round(total_geral, 2)
    }])

    df = pd.concat([df, linha_total], ignore_index=True)

    # ------------ #
    # CALCULAR BDI #
    # ------------ #

    bdi_str = entrada_bdi.get().replace(",", ".")

    try:
        bdi = float(bdi_str) / 100
    except:
        mostrar_feedback("BDI inválido.", "red")
        return

    valor_bdi = total_geral * bdi

    linha_bdi = pd.DataFrame([{
        "Código SINAPI": "",
        "Descrição do item": f"Total do BDI ({entrada_bdi.get()}%)",
        "Unid.": "",
        "Qtd.": "",
        "Valor Unit.": "",
        "Total s/ BDI": round(valor_bdi, 2)
    }])

    df = pd.concat([df, linha_bdi], ignore_index=True)

    # --------- #
    # EVENTUAIS #
    # --------- #

    base_eventuais = total_geral + valor_bdi
    valor_eventuais = base_eventuais * 0.10 if var_eventuais.get() else 0

    if var_eventuais.get():
        linha_eventuais = pd.DataFrame([{
            "Código SINAPI": "",
            "Descrição do item": "Eventuais (10%)",
            "Unid.": "",
            "Qtd.": "",
            "Valor Unit.": "",
            "Total s/ BDI": round(valor_eventuais, 2)
        }])

        df = pd.concat([df, linha_eventuais], ignore_index=True)

    # ------- #
    # ALUGUEL #
    # ------- #

    aluguel_str = entrada_aluguel.get().replace(",", ".")

    try:
        aluguel = float(aluguel_str)
    except:
        mostrar_feedback("Valor de aluguel inválido.", "red")
        return

    linha_aluguel = pd.DataFrame([{
        "Código SINAPI": "",
        "Descrição do item": "Aluguel (1 mês)",
        "Unid.": "",
        "Qtd.": "",
        "Valor Unit.": "",
        "Total s/ BDI": round(aluguel, 2)
    }])

    df = pd.concat([df, linha_aluguel], ignore_index=True)

    # ----------- #
    # TOTAL FINAL #
    # ----------- #

    total_final = total_geral + valor_bdi + valor_eventuais + aluguel

    linha_total_final = pd.DataFrame([{
        "Código SINAPI": "",
        "Descrição do item": "TOTAL GERAL",
        "Unid.": "",
        "Qtd.": "",
        "Valor Unit.": "",
        "Total s/ BDI": round(total_final, 2)
    }])

    df = pd.concat([df, linha_total_final], ignore_index=True)

    df = df[["Código SINAPI", "Descrição do item", "Unid.", "Qtd.", "Valor Unit.", "Total s/ BDI"]]

    nome_proprietario = nome_arquivo(entrada_proprietario.get())
    nome_base = f"orcamento_reparos_{nome_proprietario}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    pasta_downloads = os.path.join(os.path.expanduser("~"), "Downloads")

    if not os.path.isdir(pasta_downloads):
        pasta_downloads = os.getcwd()

    arquivo = filedialog.asksaveasfilename(
        title="Salvar orçamento como:",
        initialdir=pasta_downloads,
        initialfile=nome_base,
        defaultextension=".xlsx",
        filetypes=[("Planilha Excel", "*.xlsx")]
    )

    if not arquivo:
        mostrar_feedback("Geração de orçamento cancelada.", "orange", temporario=False)
        return

    try:
        df.to_excel(arquivo, index=False)
    except PermissionError:
        mostrar_feedback(
            "Feche a planilha antes de gerar novamente.",
            "red",
            temporario=False
        )
        return

    mostrar_feedback("Orçamento gerado com sucesso!", "dark green", temporario=False)
    
# ---------------------------- #
# FORMATAR PLANILHA            #
# ---------------------------- #

    wb = load_workbook(arquivo)
    ws = wb.active

    fonte_cabecalho = Font(bold=True)

    fundo_cabecalho = PatternFill(
        start_color="006699",
        end_color="006699",
        fill_type="solid"
    )

    fundo_anomalia = PatternFill(
        start_color='D0CECE',
        end_color='D0CECE',
        fill_type='solid'
    )
    
    fundo_totais = PatternFill(
        start_color="F2F2F2",
        end_color="F2F2F2",
        fill_type="solid"
    )

    fundo_total_final = PatternFill(
        start_color="006699",
        end_color="006699",
        fill_type="solid"
    )
    
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6):
        for cell in row:
            cell.border = borda
    
    # Linha de título logo acima do cabeçalho original
    ws.insert_rows(1)
    nome_titulo = entrada_proprietario.get().strip() or ""
    ws.merge_cells("A1:F1")
    titulo_base = "ORÇAMENTO DE REPAROS DE VÍCIOS CONSTRUTIVOS"
    ws["A1"] = (
        f"{titulo_base} - {nome_titulo.upper()}" if nome_titulo else titulo_base
    )
    ws.row_dimensions[1].height = 24.75

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    # larguras das colunas
    ws.column_dimensions["A"].width = 7.24
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 5.96

    # largura da coluna de Qtd.
    max_length = 0

    for col in ["D"]:

        for cell in ws[col]:

            if cell.value is not None:

                comprimento = len(str(cell.value))

                if comprimento > max_length:
                    max_length = comprimento

    ws.column_dimensions["D"].width = max_length + 1

    # largura da coluna de Total s/ BDI
    max_length = 0

    for col in ["F"]:

        for cell in ws[col]:

            if cell.value is not None:

                comprimento = len(str(cell.value))

                if comprimento > max_length:
                    max_length = comprimento

    ws.column_dimensions["F"].width = max_length

    # centralizar e quebrar texto do cabeçalho
    # título (linha 1)
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = fundo_cabecalho
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # cabeçalho da tabela (linha 2)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fundo_cabecalho
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
    
    colunas_centro = ["A", "C", "D", "E", "F"]

    for col in colunas_centro:

        for cell in ws[col]:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):

        for cell in row:

            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal if cell.alignment else "center",
                vertical="center",
                wrap_text=True
            )
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):

        codigo = row[0].value
        descricao = row[1].value

        # --------------------------- #
        # TOTAL SEM BDI, BDI, ALUGUEL #
        # --------------------------- #
        descricao_norm = normalizar_texto(descricao)

        if descricao_norm and (
            descricao_norm == "TOTAL SEM BDI" or
            descricao_norm.startswith("TOTAL DO BDI") or
            descricao_norm.startswith("EVENTUAIS") or
            descricao_norm == "ALUGUEL (1 MES)"
        ):
            linha_idx = row[0].row
            descricao_formatada = str(descricao).strip()

            if descricao_norm == "TOTAL SEM BDI":
                descricao_formatada = "Total sem BDI"
            elif descricao_norm == "ALUGUEL (1 MES)":
                descricao_formatada = "Aluguel (1 mês)"
            elif descricao_norm.startswith("TOTAL DO BDI"):
                descricao_formatada = f"Total do BDI ({entrada_bdi.get()}%)"

            # O texto original está na coluna B. Após mesclar, precisa ficar em A
            ws.cell(row=linha_idx, column=1, value=descricao_formatada)
            for col_idx in range(2, 6):
                ws.cell(row=linha_idx, column=col_idx, value=None)

            # Mesclar rótulo em A:E, como no modelo de referência.
            ws.merge_cells(start_row=linha_idx, start_column=1, end_row=linha_idx, end_column=5)

            for col_idx in range(1, 7):
                cell = ws.cell(row=linha_idx, column=col_idx)
                cell.fill = fundo_totais
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(
                    horizontal="right" if col_idx == 1 else "center",
                    vertical="center",
                    wrap_text=True
                )

        # ----------- #
        # TOTAL FINAL #
        # ----------- #
        elif descricao == "TOTAL GERAL":
            linha_idx = row[0].row
            descricao_formatada = "Total do Orçamento"

            ws.cell(row=linha_idx, column=1, value=descricao_formatada)
            for col_idx in range(2, 6):
                ws.cell(row=linha_idx, column=col_idx, value=None)

            ws.merge_cells(start_row=linha_idx, start_column=1, end_row=linha_idx, end_column=5)

            for col_idx in range(1, 7):
                cell = ws.cell(row=linha_idx, column=col_idx)
                cell.fill = fundo_total_final
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.alignment = Alignment(
                    horizontal="right" if col_idx == 1 else "center",
                    vertical="center",
                    wrap_text=True
                )

        # --------------------- #
        # SUBTÍTULO DE ANOMALIA #
        # --------------------- #
        elif (codigo in ("", None)) and descricao not in ("", None):

            for cell in row:
                cell.fill = fundo_anomalia
                cell.font = Font(bold=True, size=12)

    linha_nota_sinapi = ws.max_row + 1
    estado_planilha = combo_estado.get().strip()
    sufixo_referencia = (
        f"{estado_planilha} {sinapi_referencia_rotulo}"
        if estado_planilha
        else sinapi_referencia_rotulo
    )
    texto_nota = f"Base de preços: SINAPI — referência: {sufixo_referencia}"
    ws.merge_cells(
        start_row=linha_nota_sinapi, start_column=1,
        end_row=linha_nota_sinapi, end_column=6,
    )
    celula_nota = ws.cell(row=linha_nota_sinapi, column=1, value=texto_nota)
    celula_nota.font = Font(size=9, italic=True, color="444444")
    celula_nota.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb.save(arquivo)
    os.startfile(arquivo)

# ---------------------------- #
# BOTÃO GERAR ORÇAMENTO        #
# ---------------------------- #

botao_gerar = tk.Button(
    frame_principal,
    text="Gerar Orçamento",
    font=("Arial", 11, "bold"),
    padx=12,
    pady=6,
    #   # cores do botão
    #   bg="#006699",
    #   fg="white",
    #   activebackground="#00557a",
    #   activeforeground="white",
    #   relief="flat",
    #   bd=0,
    #   cursor="hand2",

    command=gerar_orcamento
)

botao_gerar.pack(pady=(5,10))

entrada_proprietario.focus()

janela.mainloop()