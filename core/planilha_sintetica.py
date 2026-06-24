"""Gera planilha intermediária no layout esperado pelo formatador de modelos.

O ORC monta essa estrutura a partir do Orçamento Customizado antes de aplicar
os modelos 1, 2 e 3. O layout é compatível com exportações do sistema i9
(externo ao ORC), mas aqui a planilha é sempre gerada internamente.
"""

from __future__ import annotations

import openpyxl

from core.composicoes_proprias import custo_composicao_propria_item
from core.formatador_sinapi.comum import planilha_ativa, valor_em_extenso
from core.orcamento_customizado import (
    TIPO_COMPOSICAO_PROPRIA,
    TIPO_SINAPI,
    custo_unitario_com_bdi,
)

LINHA_CABECALHO = 13
COLUNAS = (
    "Item",
    "Tipo",
    "Banco",
    "Código",
    "Descrição",
    "Un.",
    "Qtd.",
    "Preço Unit",
    "Preço com BDI",
    "Total sem BDI",
    "Total",
)


def _formatar_bdi_cabecalho(valor: float) -> str:
    return f"{float(valor):.2f}".replace(".", ",")


def _formatar_referencia_bancos(estado: str, referencia_rotulo: str) -> str:
    mes_ano = referencia_rotulo.strip()
    if "/" in mes_ano:
        mes, ano = mes_ano.split("/", 1)
        try:
            mes_ano = f"{int(mes)}/{ano}"
        except ValueError:
            pass
    return f"Bancos:\nSINAPI: {estado} {mes_ano}"


def _preencher_cabecalho(ws, orcamento, estado: str, referencia_rotulo: str) -> None:
    nome_obra = (orcamento.nome or "Orçamento customizado").strip()
    bdi = _formatar_bdi_cabecalho(orcamento.bdi_percent)

    ws.cell(row=2, column=1, value="Número: 1")
    ws.cell(row=4, column=1, value=f"Obra: {nome_obra}")
    ws.cell(row=5, column=1, value="Orçamento: CUSTOMIZADO")
    ws.cell(row=6, column=1, value="Cliente: ")
    ws.cell(row=6, column=5, value="Planilha intermediária ORC")
    ws.cell(row=7, column=1, value=f"BDI Padrão: {bdi}%")
    ws.cell(row=8, column=1, value=_formatar_referencia_bancos(estado, referencia_rotulo))

    for col, titulo in enumerate(COLUNAS, start=1):
        ws.cell(row=LINHA_CABECALHO, column=col, value=titulo)


def _dados_item(item, catalogo, sinapi, estado: str, bdi: float):
    if item["tipo"] == TIPO_SINAPI:
        preco_s = float(item["custo_unitario"])
        banco = "SINAPI"
        tipo = "Composição"
        codigo = str(item["codigo"]).strip()
        descricao = str(item.get("descricao", "")).strip()
        unidade = str(item.get("unidade", "")).strip()
    else:
        preco_s, _ = custo_composicao_propria_item(item, catalogo, sinapi, estado)
        preco_s = float(preco_s)
        banco = "Próprio"
        tipo = "Composição"
        codigo = str(item.get("codigo", "")).strip()
        descricao = str(item.get("nome", "")).strip()
        unidade = str(item.get("unidade", "")).strip()

    quantidade = float(item["quantidade"])
    preco_c = custo_unitario_com_bdi(preco_s, bdi)
    total_s = round(preco_s * quantidade, 2)
    total_c = round(preco_c * quantidade, 2)
    return {
        "tipo": tipo,
        "banco": banco,
        "codigo": codigo,
        "descricao": descricao,
        "unidade": unidade,
        "quantidade": quantidade,
        "preco_s": round(preco_s, 2),
        "preco_c": round(preco_c, 2),
        "total_s": total_s,
        "total_c": total_c,
    }


def gerar_planilha_sintetica(
    caminho_saida: str,
    orcamento,
    catalogo,
    sinapi,
    estado: str,
    referencia_rotulo: str,
) -> str:
    wb = openpyxl.Workbook()
    ws = planilha_ativa(wb)
    ws.title = "Orçamento"

    _preencher_cabecalho(ws, orcamento, estado, referencia_rotulo)

    linha = LINHA_CABECALHO + 1
    bdi = float(orcamento.bdi_percent)
    total_sem_bdi = 0.0

    grupos_com_itens = [g for g in orcamento.grupos if g.get("itens")]
    for indice_grupo, grupo in enumerate(grupos_com_itens):
        subtotal_grupo_s = 0.0
        subtotal_grupo_c = 0.0
        rotulo_grupo = f" {indice_grupo}"

        itens_dados = [
            _dados_item(item, catalogo, sinapi, estado, bdi) for item in grupo["itens"]
        ]
        for dados in itens_dados:
            subtotal_grupo_s += dados["total_s"]
            subtotal_grupo_c += dados["total_c"]

        subtotal_grupo_s = round(subtotal_grupo_s, 2)
        subtotal_grupo_c = round(subtotal_grupo_c, 2)
        total_sem_bdi += subtotal_grupo_s

        ws.cell(row=linha, column=1, value=rotulo_grupo)
        ws.cell(row=linha, column=5, value=grupo["nome"].strip().upper())
        ws.cell(row=linha, column=6, value=valor_em_extenso(subtotal_grupo_c))
        ws.cell(row=linha, column=11, value=subtotal_grupo_c)
        linha += 1

        for indice_item, dados in enumerate(itens_dados, start=1):
            ws.cell(row=linha, column=1, value=f" {indice_grupo}.{indice_item}")
            ws.cell(row=linha, column=2, value=dados["tipo"])
            ws.cell(row=linha, column=3, value=dados["banco"])
            ws.cell(row=linha, column=4, value=dados["codigo"])
            ws.cell(row=linha, column=5, value=dados["descricao"])
            ws.cell(row=linha, column=6, value=dados["unidade"])
            ws.cell(row=linha, column=7, value=dados["quantidade"])
            ws.cell(row=linha, column=8, value=dados["preco_s"])
            ws.cell(row=linha, column=9, value=dados["preco_c"])
            ws.cell(row=linha, column=10, value=dados["total_s"])
            ws.cell(row=linha, column=11, value=dados["total_c"])
            linha += 1

    total_sem_bdi = round(total_sem_bdi, 2)
    valor_bdi = round(total_sem_bdi * bdi / 100, 2)
    total_final = round(total_sem_bdi + valor_bdi, 2)

    linha += 1
    ws.cell(row=linha, column=9, value="Total sem BDI")
    ws.cell(row=linha, column=11, value=total_sem_bdi)
    linha += 1
    ws.cell(row=linha, column=9, value="Total do BDI")
    ws.cell(row=linha, column=11, value=valor_bdi)
    linha += 1
    ws.cell(row=linha, column=9, value="Total")
    ws.cell(row=linha, column=11, value=total_final)

    wb.save(caminho_saida)
    return caminho_saida
