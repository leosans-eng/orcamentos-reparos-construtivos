"""Formatação da planilha — Modelo 1 (Atualização)."""

from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.formatador_sinapi.comum import (
    FORMATO_MOEDA,
    aplicar_borda_contorno,
    encontrar_linha_dados_start,
    extrair_bdi_rotulo,
    extrair_nome_obra,
    extrair_referencia_sinapi,
    extrair_totais_finais,
    gerar_caminho_saida,
    atribuir_valor,
    planilha_ativa,
    valor_em_extenso,
)
from core.formatador_sinapi.types import Modelo, ResultadoFormatacao


def formatar_modelo1(
    caminho_origem_xlsx: str,
    *,
    caminho_saida: str | None = None,
    diretorio_saida: str | None = None,
) -> ResultadoFormatacao:
    if not os.path.isfile(caminho_origem_xlsx):
        raise FileNotFoundError(
            f"O arquivo '{caminho_origem_xlsx}' não foi encontrado."
        )

    wb_origem = openpyxl.load_workbook(caminho_origem_xlsx)
    ws_origem = planilha_ativa(wb_origem)

    nome_obra = extrair_nome_obra(ws_origem)
    referencia_sinapi = extrair_referencia_sinapi(ws_origem)
    caminho_saida_xlsx = gerar_caminho_saida(
        caminho_origem_xlsx,
        Modelo.ATUALIZACAO,
        nome_obra,
        caminho_saida=caminho_saida,
        diretorio_saida=diretorio_saida,
    )

    wb_destino = openpyxl.Workbook()
    ws_destino = planilha_ativa(wb_destino)
    ws_destino.title = "Orçamento Formatado"
    ws_destino.views.sheetView[0].showGridLines = True

    font_cabecalho = Font(name="Calibri", size=11, bold=True, color="000000")
    fill_cabecalho = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    font_secao = Font(name="Calibri", size=12, bold=False, color="000000")
    fill_secao = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    font_corpo = Font(name="Calibri", size=10, bold=False, color="000000")
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_direita = Alignment(horizontal="right", vertical="center")
    borda_fina = Side(style="thin", color="EDEDED")
    grade_celula = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    colunas_modelo = [
        "Item", "Código Sinapi", "Descrição", "Un.", "Qtd.",
        "Preço s/ BDI", "Preço c/ BDI", "Total s/ BDI", "Total c/ BDI",
    ]
    ws_destino.append(colunas_modelo)
    ws_destino.row_dimensions[1].height = 32

    for col_idx, _nome_coluna in enumerate(colunas_modelo, 1):
        celula = ws_destino.cell(row=1, column=col_idx)
        celula.font = font_cabecalho
        celula.fill = fill_cabecalho
        celula.alignment = align_centro
        celula.border = grade_celula

    linha_dados_start = encontrar_linha_dados_start(ws_origem)

    for row_idx in range(linha_dados_start, ws_origem.max_row + 1):
        item_val = ws_origem.cell(row=row_idx, column=1).value
        texto_extenso = str(ws_origem.cell(row=row_idx, column=4).value or "").strip()

        if (item_val is None or str(item_val).strip() == "") and texto_extenso == "":
            continue

        item = str(item_val).strip()
        banco = str(ws_origem.cell(row=row_idx, column=3).value or "").strip()
        codigo_original = str(ws_origem.cell(row=row_idx, column=4).value or "").strip()
        codigo = "Comp. SINAPI" if banco.lower() == "próprio" else codigo_original
        descricao = str(ws_origem.cell(row=row_idx, column=5).value or "").strip()

        limite_linha = 85
        num_linhas = max(1, (len(descricao) // limite_linha) + 1)
        altura_linha = num_linhas * 15
        if codigo == "Comp. SINAPI":
            altura_linha = max(altura_linha, 30)

        qtd = ws_origem.cell(row=row_idx, column=7).value
        preco_s_bdi = ws_origem.cell(row=row_idx, column=8).value
        preco_c_bdi = ws_origem.cell(row=row_idx, column=9).value
        total_s_bdi = ws_origem.cell(row=row_idx, column=10).value
        total_c_bdi = ws_origem.cell(row=row_idx, column=11).value

        num_linha_atual = ws_destino.max_row + 1
        ws_destino.row_dimensions[num_linha_atual].height = altura_linha

        if texto_extenso and banco == "" and descricao == "":
            ws_destino.append(["", "", "", texto_extenso, "", "", "", "", ""])
            linha_extenso = ws_destino.max_row
            ws_destino.merge_cells(
                start_row=linha_extenso, start_column=4, end_row=linha_extenso, end_column=8
            )
            celula_extenso = ws_destino.cell(row=linha_extenso, column=4)
            celula_extenso.font = Font(name="Calibri", size=9, bold=False)
            celula_extenso.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            linhas_extenso = max(1, (len(texto_extenso) // 90) + 1)
            ws_destino.row_dimensions[linha_extenso].height = linhas_extenso * 14
            continue

        if codigo_original == "":
            valor_extenso = valor_em_extenso(total_c_bdi)
            ws_destino.append([item, "", descricao, valor_extenso, "", "", "", "", total_c_bdi])
            ws_destino.row_dimensions[num_linha_atual].height = max(altura_linha, 26)
            ws_destino.merge_cells(
                start_row=num_linha_atual, start_column=4, end_row=num_linha_atual, end_column=8
            )
            for c_idx in range(1, 10):
                celula = ws_destino.cell(row=num_linha_atual, column=c_idx)
                celula.fill = fill_secao
                celula.border = grade_celula
                if c_idx == 4:
                    celula.font = Font(name="Calibri", size=9, bold=False, color="000000")
                else:
                    celula.font = font_secao
                if c_idx in [1, 2]:
                    celula.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                elif c_idx in [3, 4]:
                    celula.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                elif c_idx == 9:
                    celula.alignment = align_direita
                    celula.number_format = FORMATO_MOEDA
        else:
            un = str(ws_origem.cell(row=row_idx, column=6).value or "").strip()
            ws_destino.append(
                [item, codigo, descricao, un, qtd, preco_s_bdi, preco_c_bdi, total_s_bdi, total_c_bdi]
            )
            for c_idx in range(1, 10):
                celula = ws_destino.cell(row=num_linha_atual, column=c_idx)
                celula.font = font_corpo
                celula.border = grade_celula
                if c_idx in [1, 2]:
                    celula.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                elif c_idx in [3, 4]:
                    celula.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                elif c_idx == 9:
                    celula.alignment = align_direita
                    celula.number_format = FORMATO_MOEDA
                else:
                    celula.alignment = Alignment(horizontal="general", vertical="center")
                    if c_idx in [6, 7, 8]:
                        celula.number_format = FORMATO_MOEDA

    totais_finais = extrair_totais_finais(ws_origem)
    tabela_final = ws_destino.max_row
    col_label_totais = 11
    col_valor_totais = 12
    if totais_finais:
        bdi_rotulo = extrair_bdi_rotulo(ws_origem)
        labels_default = [
            "Total sem BDI",
            f"Total do BDI ({bdi_rotulo}%)",
            "Total do Orçamento",
        ]
        for index, tot in enumerate(totais_finais):
            row = tabela_final + 1 + index
            label = labels_default[index] if index < len(labels_default) else tot["label"]
            valor = tot["valor_k"] if tot["valor_k"] is not None else tot["valor_i"]
            atribuir_valor(ws_destino.cell(row=row, column=col_label_totais), label)
            atribuir_valor(ws_destino.cell(row=row, column=col_valor_totais), valor)

            cel_label = ws_destino.cell(row=row, column=col_label_totais)
            cel_label.font = font_secao
            cel_label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cel_label.border = grade_celula

            cel_val = ws_destino.cell(row=row, column=col_valor_totais)
            cel_val.font = Font(name="Calibri", size=12, bold=False, color="000000")
            cel_val.alignment = align_direita
            cel_val.border = grade_celula
            if isinstance(valor, (int, float)):
                cel_val.number_format = FORMATO_MOEDA

            if index == 2:
                cel_label.font = Font(name="Calibri", size=14, bold=True, color="1F2610")
                cel_label.fill = fill_secao
                cel_val.font = Font(name="Calibri", size=14, bold=True, color="1F2610")
                cel_val.fill = fill_secao

        linha_totais_inicio = tabela_final + 1
        linha_totais_fim = tabela_final + len(totais_finais)
        aplicar_borda_contorno(
            ws_destino, linha_totais_inicio, col_label_totais, linha_totais_fim, col_valor_totais
        )

        valor_total_orcamento = (
            totais_finais[2]["valor_k"]
            if len(totais_finais) > 2 and totais_finais[2]["valor_k"] is not None
            else totais_finais[2]["valor_i"]
        )
        texto_extenso_total = valor_em_extenso(valor_total_orcamento)
        linha_extenso_total = linha_totais_fim + 2
        atribuir_valor(
            ws_destino.cell(row=linha_extenso_total, column=col_label_totais),
            texto_extenso_total,
        )

        cel_extenso = ws_destino.cell(row=linha_extenso_total, column=col_label_totais)
        cel_extenso.font = Font(name="Calibri", size=10, bold=False, color="000000")
        cel_extenso.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cel_extenso.border = grade_celula

        if referencia_sinapi:
            linha_referencia_sinapi = linha_extenso_total + 1
            atribuir_valor(
                ws_destino.cell(row=linha_referencia_sinapi, column=col_label_totais),
                referencia_sinapi,
            )
            cel_referencia = ws_destino.cell(row=linha_referencia_sinapi, column=col_label_totais)
            cel_referencia.font = Font(name="Calibri", size=10, bold=True, color="000000")
            cel_referencia.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cel_referencia.border = grade_celula

    ws_destino.column_dimensions["A"].width = 5
    ws_destino.column_dimensions["B"].width = 8
    ws_destino.column_dimensions["C"].width = 80
    ws_destino.column_dimensions["D"].width = 5
    ws_destino.column_dimensions["E"].width = 7
    ws_destino.column_dimensions["F"].width = 12
    ws_destino.column_dimensions["G"].width = 12
    ws_destino.column_dimensions["H"].width = 14
    ws_destino.column_dimensions["I"].width = 18
    ws_destino.column_dimensions["K"].width = 25
    ws_destino.column_dimensions["L"].width = 22

    aplicar_borda_contorno(ws_destino, 1, 1, tabela_final, 9)
    wb_destino.save(caminho_saida_xlsx)

    return ResultadoFormatacao(
        modelo=Modelo.ATUALIZACAO,
        caminho_origem=os.path.abspath(caminho_origem_xlsx),
        caminho_excel=os.path.abspath(caminho_saida_xlsx),
        nome_obra=nome_obra,
        referencia_sinapi=referencia_sinapi,
    )
